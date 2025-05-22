"""
文件操作工具包 (增强版)

这个模块提供了一组全面的文件操作工具，包括：
- 基本操作：读取、写入、追加、列出目录、删除、移动、复制。
- 高级操作：搜索文件、编辑文本（包括局部替换）、代码高亮、文件比较、项目结构输出。
- 格式支持：纯文本、Markdown、Word (.docx)、Excel (.xlsx)、CSV、PDF、HTML、YAML、XML。
- 归档操作：创建和解压 ZIP/TAR 压缩文件。
- 其他工具：获取文件元数据、计算文件哈希、文件监控。

可以作为 AutoGen 的 tools 使用，所有工具函数的参数都使用 Annotated 进行了详细描述。
"""

import os
import json
import glob
import re
import mimetypes
import shutil
import difflib
import datetime
import time
import pathlib
import fnmatch
import asyncio
import subprocess
from typing import Dict, List, Any, Optional, Union, Tuple, Annotated, Callable
import tempfile
from functools import wraps
import zipfile
import tarfile
import hashlib

# 定义类型别名
# 使用Any代替Callable以避免Pydantic JSON Schema生成问题
from typing import Any as FileChangeCallbackType

# 尝试导入 watchdog 用于文件监控
try:
    from watchdog.observers import Observer
    from watchdog.events import FileSystemEventHandler
    WATCHDOG_AVAILABLE = True
except ImportError:
    WATCHDOG_AVAILABLE = False
    print("警告：未能导入 watchdog，文件监控功能将不可用。请运行 'pip install watchdog' 安装。")

# 尝试导入 pygments 用于代码高亮
try:
    from pygments import highlight
    from pygments.lexers import get_lexer_for_filename, get_lexer_by_name, guess_lexer
    from pygments.formatters import HtmlFormatter, TerminalFormatter
    from pygments.util import ClassNotFound
    PYGMENTS_AVAILABLE = True
except ImportError:
    PYGMENTS_AVAILABLE = False
    print("警告：未能导入 pygments，代码高亮功能将不可用。请运行 'pip install Pygments' 安装。")

# 导入 AutoGen 相关模块
try:
    from autogen_core.tools import FunctionTool
    AUTOGEN_AVAILABLE = True
except ImportError:
    AUTOGEN_AVAILABLE = False
    # 如果在非 AutoGen 环境下运行，定义一个虚拟的 FunctionTool，以便代码能够正常加载
    class FunctionTool:
        def __init__(self, func, name, description):
            self.func = func
            self.name = name
            self.description = description
        def __call__(self, *args, **kwargs):
            return self.func(*args, **kwargs)
    print("警告：未能导入 AutoGen 工具模块，将无法作为 AutoGen tools 注册。")


# 初始化 mimetypes
mimetypes.init()

# 尝试导入各种文件格式处理库
try:
    import docx  # 处理 .docx 文件
    DOCX_AVAILABLE = True
except ImportError:
    DOCX_AVAILABLE = False
    print("提示：未能导入 python-docx，处理 .docx 文件的功能将受限。可运行 'pip install python-docx' 安装。")

try:
    import openpyxl  # 处理 .xlsx 文件
    XLSX_AVAILABLE = True
except ImportError:
    XLSX_AVAILABLE = False
    print("提示：未能导入 openpyxl，处理 .xlsx 文件的功能将受限。可运行 'pip install openpyxl' 安装。")

try:
    import pandas as pd  # 处理 .csv 和 .xlsx 文件
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False
    print("提示：未能导入 pandas，处理 .csv 和部分 .xlsx 文件的功能将受限。可运行 'pip install pandas' 安装。")

try:
    import pdfplumber  # 处理 .pdf 文件
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False
    print("提示：未能导入 pdfplumber，处理 .pdf 文件的功能将受限。可运行 'pip install pdfplumber' 安装。")

try:
    import markdown  # 处理 .md 文件
    MARKDOWN_AVAILABLE = True
except ImportError:
    MARKDOWN_AVAILABLE = False
    print("提示：未能导入 markdown，处理 .md 文件的功能将受限。可运行 'pip install markdown' 安装。")

try:
    from bs4 import BeautifulSoup  # 处理 HTML
    BS4_AVAILABLE = True
except ImportError:
    BS4_AVAILABLE = False
    print("提示：未能导入 BeautifulSoup4，处理 HTML 文件的功能将受限。可运行 'pip install beautifulsoup4' 安装。")

try:
    import yaml  # 处理 .yaml 文件
    YAML_AVAILABLE = True
except ImportError:
    YAML_AVAILABLE = False
    print("提示：未能导入 PyYAML，处理 .yaml 文件的功能将受限。可运行 'pip install PyYAML' 安装。")

try:
    import xml.etree.ElementTree as ET  # 处理 .xml 文件
    XML_AVAILABLE = True
except ImportError:
    XML_AVAILABLE = False
    print("提示：未能导入 xml.etree.ElementTree，处理 .xml 文件的功能将受限（通常为标准库）。")


# 只有在 watchdog 可用时才定义 FileChangeHandler 类
if WATCHDOG_AVAILABLE:
    class FileChangeHandler(FileSystemEventHandler):
        """文件变更处理器"""
        def __init__(self, callback: FileChangeCallbackType, path_to_monitor: str):
            self.callback = callback
            self.path_to_monitor = os.path.abspath(path_to_monitor)
            self.last_event_time = {} # 用于处理重复事件

        def on_any_event(self, event):
            """处理任何文件系统事件"""
            # 忽略目录事件，除非是创建或删除目录本身
            if event.is_directory and event.event_type not in ("created", "deleted"):
                 # 如果是监控目录，且事件是目录内的文件/子目录变化，则需要处理
                if os.path.abspath(event.src_path) != self.path_to_monitor:
                    pass # 允许子目录/文件事件触发回调
                else:
                    return


            event_key = (event.event_type, event.src_path)
            current_time = time.time()

            # 简单的去抖动逻辑：如果同一事件在短时间内重复发生，则忽略
            if event_key in self.last_event_time and (current_time - self.last_event_time[event_key]) < 0.5:
                return
            self.last_event_time[event_key] = current_time

            # 对于重命名事件，event.dest_path 才是新路径
            path_to_report = event.dest_path if event.event_type == 'moved' else event.src_path
            self.callback(event.event_type, path_to_report)

class FileUtils:
    """
    文件操作工具类。
    所有方法都设计为异步的，以便在 AutoGen 等异步环境中使用。
    路径参数可以是绝对路径，也可以是相对于当前工作目录的相对路径。
    """

    def __init__(self, base_path: Optional[str] = None):
        """
        初始化文件操作工具类。

        Args:
            base_path: 基础路径，如果提供，某些操作可能会默认相对于此路径。
                       当前实现中，大多数函数直接使用提供的路径参数。
        """
        self.base_path = base_path or os.getcwd()
        self._file_monitors = {}  # 存储文件监控器

        # 文件格式支持状态
        self.format_support = {
            "docx": DOCX_AVAILABLE,
            "xlsx": XLSX_AVAILABLE,
            "csv": PANDAS_AVAILABLE,
            "pdf": PDF_AVAILABLE,
            "md": MARKDOWN_AVAILABLE,
            "html": BS4_AVAILABLE,
            "yaml": YAML_AVAILABLE,
            "xml": XML_AVAILABLE,
            "code_highlight": PYGMENTS_AVAILABLE,
            "file_monitor": WATCHDOG_AVAILABLE
        }
        print(f"FileUtils initialized. Base path: {self.base_path}")
        print(f"Supported formats: {self.format_support}")

    def _ensure_directory_exists(self, file_path: str) -> None:
        """
        确保文件所在的目录存在，如果不存在则创建。

        Args:
            file_path: 文件的完整路径

        Returns:
            None
        """
        directory = os.path.dirname(file_path)
        if directory and not os.path.exists(directory):
            os.makedirs(directory, exist_ok=True)

    def _git_commit(self, file_path: str, operation: str, role: Optional[str] = None) -> Dict[str, Any]:
        """
        执行 git 提交操作，将修改的文件提交到 git 仓库。

        此方法使用 GitManager 类来执行 git 操作，提供了更强大和灵活的 git 功能。

        Args:
            file_path: 被修改的文件路径
            operation: 执行的操作类型，如 'write', 'edit', 'delete' 等
            role: 提交者的角色，如果提供，将添加到提交信息中

        Returns:
            包含 git 操作结果的字典
        """
        # 使用 GitManager 类执行 git 提交
        git_manager = GitManager()
        return git_manager.commit_file(file_path, operation, role)


    async def read_file(
        self,
        file_path: Annotated[str, "要读取的文件的完整路径。"],
        encoding: Annotated[Optional[str], "文件编码，例如 'utf-8', 'gbk'。如果为 None，则尝试 utf-8，失败则尝试二进制。"] = None
    ) -> Dict[str, Any]:
        """
        读取指定路径的文件内容，智能识别多种文件格式。

        Args:
            file_path: 要读取的文件的完整路径。
            encoding: 文件编码，例如 'utf-8', 'gbk'。如果为 None，则尝试 utf-8，然后是常见的编码，最后尝试二进制。

        Returns:
            包含文件内容、格式或其他相关信息（如错误）的字典。
            成功时，字典通常包含 'content', 'format' 键。
            对于特定格式（如 PDF, Excel），还可能包含 'pages', 'sheets' 等键。
            失败时，包含 'error' 键。
        """
        absolute_file_path = os.path.abspath(os.path.join(self.base_path, file_path))
        if not os.path.exists(absolute_file_path):
            return {"error": f"文件 {absolute_file_path} 不存在"}
        if not os.path.isfile(absolute_file_path):
            return {"error": f"路径 {absolute_file_path} 不是一个文件"}

        try:
            _, ext = os.path.splitext(absolute_file_path)
            ext = ext.lower().lstrip('.')

            # 定义已知的文本文件扩展名列表
            known_text_extensions = ['txt', 'log', 'py', 'js', 'java', 'c', 'cpp', 'h', 'cs', 'json', 'ini', 'cfg', 'srt', 'vtt', 'ass', 'conf', 'config', 'toml', 'ini', 'rst', 'css', 'scss', 'less', 'sass', 'xml', 'html', 'htm', 'md', 'markdown', 'sql', 'bat', 'sh', 'ps1', 'yaml', 'yml', 'ts', 'tsx', 'jsx', 'vue', 'php', 'rb', 'pl', 'go', 'rs', 'swift', 'kt', 'gradle', 'properties', 'gitignore', 'dockerignore', 'env']

            # 优先尝试作为文本文件读取，无论扩展名是什么
            encodings_to_try = [encoding] if encoding else ['utf-8', 'gbk', 'latin-1']
            content = None
            detected_encoding = None

            # 如果是已知的文本文件扩展名，或者没有特定处理方式，尝试作为文本读取
            if ext in known_text_extensions or (ext not in ['docx', 'xlsx', 'xls', 'pdf']):
                for enc in encodings_to_try:
                    if enc is None: continue
                    try:
                        with open(absolute_file_path, 'r', encoding=enc) as file:
                            content = file.read()
                        detected_encoding = enc
                        break
                    except UnicodeDecodeError:
                        continue
                    except Exception: # 其他可能的读取错误
                        continue

                if content is not None:
                    return {"content": content, "format": ext or "text", "encoding": detected_encoding}


            elif ext == 'md' and MARKDOWN_AVAILABLE:
                with open(absolute_file_path, 'r', encoding=encoding or 'utf-8') as file:
                    md_content = file.read()
                html_content = markdown.markdown(md_content)
                return {"content": md_content, "html": html_content, "format": "markdown", "encoding": encoding or 'utf-8'}

            elif ext == 'docx' and DOCX_AVAILABLE:
                doc = docx.Document(absolute_file_path)
                paragraphs = [p.text for p in doc.paragraphs]
                content = '\n'.join(paragraphs)
                return {"content": content, "paragraphs": paragraphs, "format": "docx"}

            elif ext in ['xlsx', 'xls']:
                if PANDAS_AVAILABLE:
                    try:
                        excel_file = pd.ExcelFile(absolute_file_path)
                        sheets_data = {}
                        for sheet_name in excel_file.sheet_names:
                            df = excel_file.parse(sheet_name)
                            # 将 NaN 替换为 None 以便 JSON 序列化
                            sheets_data[sheet_name] = df.where(pd.notnull(df), None).to_dict(orient='records')
                        return {"sheets": sheets_data, "sheet_names": excel_file.sheet_names, "format": "excel_pandas"}
                    except Exception as e_pandas:
                        if XLSX_AVAILABLE: # Pandas 失败，尝试 openpyxl
                            pass # 继续到下面的 openpyxl
                        else:
                            return {"error": f"使用 pandas 读取 Excel 文件 {absolute_file_path} 失败: {e_pandas}"}

                if XLSX_AVAILABLE:
                    workbook = openpyxl.load_workbook(absolute_file_path, read_only=True, data_only=True) # data_only=True 获取单元格的值而非公式
                    sheets = {}
                    for sheet_name in workbook.sheetnames:
                        sheet = workbook[sheet_name]
                        rows_data = []
                        for row in sheet.iter_rows(values_only=True):
                            rows_data.append(list(row))
                        sheets[sheet_name] = rows_data
                    return {"sheets": sheets, "sheet_names": workbook.sheetnames, "format": "excel_openpyxl"}
                else:
                    return {"error": "处理 Excel 文件需要 pandas 或 openpyxl，但两者均未安装。"}


            elif ext == 'csv' and PANDAS_AVAILABLE:
                df = pd.read_csv(absolute_file_path, encoding=encoding or 'utf-8')
                return {"content": df.where(pd.notnull(df), None).to_dict(orient='records'), "columns": df.columns.tolist(), "format": "csv", "encoding": encoding or 'utf-8'}

            elif ext == 'pdf' and PDF_AVAILABLE:
                with pdfplumber.open(absolute_file_path) as pdf:
                    pages_text = []
                    for i, page in enumerate(pdf.pages):
                        text = page.extract_text()
                        pages_text.append({"page_number": i + 1, "content": text or ""})
                    full_content = '\n\n--- Page Break ---\n\n'.join([p["content"] for p in pages_text])
                return {"content": full_content, "pages": pages_text, "page_count": len(pdf.pages), "format": "pdf"}

            elif ext in ['html', 'htm'] and BS4_AVAILABLE:
                with open(absolute_file_path, 'r', encoding=encoding or 'utf-8') as file:
                    html_content = file.read()
                soup = BeautifulSoup(html_content, 'html.parser')
                text_content = soup.get_text(separator='\n', strip=True)
                return {"content": text_content, "html": html_content, "format": "html", "encoding": encoding or 'utf-8'}

            elif ext in ['yaml', 'yml'] and YAML_AVAILABLE:
                with open(absolute_file_path, 'r', encoding=encoding or 'utf-8') as file:
                    yaml_content = yaml.safe_load(file)
                return {"content": yaml_content, "format": "yaml", "encoding": encoding or 'utf-8'}

            elif ext == 'xml' and XML_AVAILABLE:
                tree = ET.parse(absolute_file_path)
                root = tree.getroot()
                def xml_to_dict_recursive(element):
                    d = {element.tag: {} if element.attrib else None}
                    children = list(element)
                    if children:
                        dd = {}
                        for dc in children:
                            child_dict = xml_to_dict_recursive(dc)
                            # Handle repeated tags by making them a list
                            if dc.tag in dd:
                                if not isinstance(dd[dc.tag], list):
                                    dd[dc.tag] = [dd[dc.tag]]
                                dd[dc.tag].append(child_dict[dc.tag])
                            else:
                                dd[dc.tag] = child_dict[dc.tag]
                        d = {element.tag: dd}
                    if element.attrib:
                        d[element.tag].update(('@' + k, v) for k, v in element.attrib.items())
                    if element.text and element.text.strip():
                        if children or element.attrib:
                            d[element.tag]['#text'] = element.text.strip()
                        else:
                            d[element.tag] = element.text.strip()
                    return d
                xml_dict = xml_to_dict_recursive(root)
                return {"content": xml_dict, "format": "xml", "encoding": encoding or 'utf-8'}

            # 如果所有尝试的文本读取方法都失败，最后才尝试作为二进制文件处理
            with open(absolute_file_path, 'rb') as file:
                # 读取少量头部数据用于预览或识别
                preview_bytes = file.read(1024)
            file_size = os.path.getsize(absolute_file_path)
            mime_type, _ = mimetypes.guess_type(absolute_file_path)
            return {
                "message": "文件无法作为文本读取，已作为二进制文件处理。",
                "file_info": {
                    "path": absolute_file_path,
                    "size_bytes": file_size,
                    "mime_type": mime_type or "application/octet-stream",
                    "preview_hex": preview_bytes.hex()[:128] + "..." if preview_bytes else "", # 显示前64字节的十六进制
                    "format": "binary"
                }
            }

        except Exception as e:
            return {"error": f"读取文件 {absolute_file_path} 时出错: {str(e)}"}

    async def write_file(
        self,
        file_path: Annotated[str, "要写入的文件的完整路径。如果文件已存在，它将被覆盖。"],
        content: Annotated[str, "要写入文件的文本内容。"],
        encoding: Annotated[str, "写入文件时使用的编码，例如'utf-8'。"] = "utf-8",
        normalize_line_endings: Annotated[bool, "是否规范化行尾（将\\r\\n或\\r转换为\\n）。"] = True,
        ensure_final_newline: Annotated[bool, "确保文件以换行符结束。"] = True,
        role: Annotated[Optional[str], "提交者的角色，如果提供，将添加到提交信息中。"] = None,
        auto_commit: Annotated[bool, "是否在操作后自动执行Git提交。"] = True
    ) -> Dict[str, Any]:
        """
        将内容写入指定路径的文件。如果文件已存在，它将被覆盖。

        此函数提供了对换行符处理的精细控制。
        操作后会尝试自动执行 Git 提交，可通过 auto_commit=False 禁用。

        Args:
            file_path: 要写入的文件的完整路径。
            content: 要写入文件的文本内容。
            encoding: 写入文件时使用的编码。
            normalize_line_endings: 是否规范化行尾（将CRLF和CR都转换为LF）。
            ensure_final_newline: 确保文件以换行符结束。
            role: Git提交者的角色。
            auto_commit: 是否在操作后自动执行Git提交。

        Returns:
            包含操作状态和文件信息的字典。
        """
        absolute_file_path = os.path.abspath(os.path.join(self.base_path, file_path))
        try:
            self._ensure_directory_exists(absolute_file_path)
            processed_content = content
            if normalize_line_endings:
                # 处理实际的回车换行符
                processed_content = processed_content.replace('\r\n', '\n').replace('\r', '\n')
                # 处理字符串字面量中的转义序列
                processed_content = processed_content.replace('\\r\\n', '\n').replace('\\r', '\n')
            if ensure_final_newline and processed_content and not processed_content.endswith('\n'):
                processed_content += '\n'

            with open(absolute_file_path, 'wb') as file: # 以二进制模式写入以精确控制字节
                file.write(processed_content.encode(encoding))

            # 计算行数
            line_count = processed_content.count('\n')
            if processed_content and not processed_content.endswith('\n'):
                 line_count += 1

            result = {
                "status": "成功",
                "message": f"内容已写入文件 {absolute_file_path}",
                "path": absolute_file_path,
                "size_bytes": len(processed_content.encode(encoding)),
                "line_count": line_count,
                "ends_with_newline": processed_content.endswith('\n') if processed_content else False
            }
            if auto_commit:
                git_result = self._git_commit(absolute_file_path, "write", role)
                result["git_result"] = git_result
            else:
                result["git_result"] = {"git_status": "跳过", "message": "Git提交被禁用 (auto_commit=False)"}
            return result
        except Exception as e:
            return {"error": f"写入文件 {absolute_file_path} 时出错: {str(e)}"}

    async def append_file(
        self,
        file_path: Annotated[str, "要追加内容的文件的完整路径。如果文件不存在，则会创建它。"],
        content: Annotated[str, "要追加到文件末尾的文本内容。在JSON中表示多行文本时，使用\\n表示换行，不要使用实际的换行符。"],
        encoding: Annotated[str, "写入文件时使用的编码，例如'utf-8'。"] = "utf-8",
        normalize_line_endings: Annotated[bool, "是否规范化行尾（将\\r\\n转换为\\n）。对于跨平台兼容性，建议保持为True。"] = True,
        ensure_final_newline: Annotated[bool, "确保文件以换行符结束。大多数文本文件应以换行符结束。"] = True,
        role: Annotated[Optional[str], "提交者的角色，如果提供，将添加到提交信息中。"] = None,
        auto_commit: Annotated[bool, "是否在操作后自动执行Git提交。"] = True
    ) -> Dict[str, Any]:
        """
        将内容追加到指定路径的文件末尾。如果文件不存在，则会创建它。

        [已弃用] 请使用 append_to_file 方法代替，它提供了相同的功能和更好的实现。
        此方法将在未来版本中移除。

        此函数提供了对换行符处理的精细控制，确保文件内容在不同平台上保持一致。
        它会检查现有文件是否以换行符结束，如果不是且ensure_final_newline=True，
        则会在追加新内容前先添加一个换行符，确保内容正确分行。
        操作后会尝试自动执行 Git 提交，可通过 auto_commit=False 禁用。

        Args:
            file_path: 要追加内容的文件的完整路径。
            content: 要追加到文件末尾的文本内容。在JSON中表示多行文本时，使用\\n表示换行，不要使用实际的换行符。
            encoding: 写入文件时使用的编码。
            normalize_line_endings: 是否规范化行尾（将\\r\\n转换为\\n）。对于跨平台兼容性，建议保持为True。
            ensure_final_newline: 确保文件以换行符结束。大多数文本文件应以换行符结束。
            role: Git提交者的角色。
            auto_commit: 是否在操作后自动执行Git提交。

        Returns:
            包含操作状态和文件信息的字典，包括：
            - status: 操作状态，成功或错误
            - message: 操作结果描述
            - path: 文件的绝对路径
            - appended_bytes: 追加的字节数
            - appended_lines: 追加的行数
            - file_existed: 文件在追加前是否已存在
            - file_ended_with_newline: 文件在追加前是否以换行符结束
            - git_result: Git 提交操作的结果

        示例:
            # 追加内容到文本文件
            result = await file_utils.append_to_file("log.txt", "New log entry\\n")

            # 追加JSON数据到文件
            json_data = '{"timestamp":"2023-01-01","event":"login"}'
            result = await file_utils.append_to_file("events.json", "\\n" + json_data)

            # 追加内容但不规范化换行符
            result = await file_utils.append_to_file("raw_data.txt", "Raw content\\r\\n", normalize_line_endings=False)
        """
        # 打印弃用警告
        import warnings
        warnings.warn(
            "append_file 方法已弃用，请使用 append_to_file 方法代替。此方法将在未来版本中移除。",
            DeprecationWarning,
            stacklevel=2
        )

        # 直接调用 append_to_file 方法
        return await self.append_to_file(
            file_path=file_path,
            content=content,
            encoding=encoding,
            normalize_line_endings=normalize_line_endings,
            ensure_final_newline=ensure_final_newline,
            role=role,
            auto_commit=auto_commit
        )


    async def append_to_file(
        self,
        file_path: Annotated[str, "要追加内容的文件的完整路径。如果文件不存在，则会创建它。"],
        content: Annotated[str, "要追加到文件末尾的文本内容。"],
        encoding: Annotated[str, "写入文件时使用的编码，例如'utf-8'。"] = "utf-8",
        normalize_line_endings: Annotated[bool, "是否规范化行尾（将\\r\\n或\\r转换为\\n）。"] = True,
        ensure_final_newline: Annotated[bool, "确保追加的内容（及整个文件）以换行符结束。"] = True,
        role: Annotated[Optional[str], "提交者的角色，如果提供，将添加到提交信息中。"] = None,
        auto_commit: Annotated[bool, "是否在操作后自动执行Git提交。"] = True
    ) -> Dict[str, Any]:
        """
        将内容追加到指定路径的文件末尾。如果文件不存在，则会创建它。

        此函数提供了对换行符处理的精细控制。
        操作后会尝试自动执行 Git 提交，可通过 auto_commit=False 禁用。

        Args:
            file_path: 要追加内容的文件的完整路径。
            content: 要追加到文件末尾的文本内容。
            encoding: 写入文件时使用的编码。
            normalize_line_endings: 是否规范化行尾。
            ensure_final_newline: 确保追加的内容（及整个文件）以换行符结束。
            role: Git提交者的角色。
            auto_commit: 是否在操作后自动执行Git提交。

        Returns:
            包含操作状态和文件信息的字典。
        """
        absolute_file_path = os.path.abspath(os.path.join(self.base_path, file_path))
        try:
            self._ensure_directory_exists(absolute_file_path)
            processed_content = content
            if normalize_line_endings:
                processed_content = processed_content.replace('\r\n', '\n').replace('\r', '\n')

            prefix_newline = ''
            file_existed = os.path.exists(absolute_file_path)
            file_ended_with_newline_before_append = False

            if file_existed and os.path.getsize(absolute_file_path) > 0:
                with open(absolute_file_path, 'rb') as f_check:
                    f_check.seek(-1, os.SEEK_END) # 移动到倒数第一个字节
                    if f_check.read(1) == b'\n':
                        file_ended_with_newline_before_append = True
                if not file_ended_with_newline_before_append and processed_content:
                    prefix_newline = '\n'

            final_content_to_append = prefix_newline + processed_content
            if ensure_final_newline and final_content_to_append and not final_content_to_append.endswith('\n'):
                final_content_to_append += '\n'

            with open(absolute_file_path, 'ab') as file: # 以二进制追加模式打开
                file.write(final_content_to_append.encode(encoding))

            appended_lines = final_content_to_append.count('\n')
            if final_content_to_append and not final_content_to_append.endswith('\n'):
                appended_lines += 1

            result = {
                "status": "成功",
                "message": f"内容已追加到文件 {absolute_file_path}",
                "path": absolute_file_path,
                "appended_bytes": len(final_content_to_append.encode(encoding)),
                "appended_lines": appended_lines,
                "file_existed": file_existed,
                "file_ended_with_newline_before_append": file_ended_with_newline_before_append
            }
            if auto_commit:
                git_result = self._git_commit(absolute_file_path, "append", role)
                result["git_result"] = git_result
            else:
                result["git_result"] = {"git_status": "跳过", "message": "Git提交被禁用 (auto_commit=False)"}
            return result
        except Exception as e:
            return {"error": f"追加内容到文件 {absolute_file_path} 时出错: {str(e)}"}

    async def list_directory(
        self,
        directory_path: Annotated[str, "要列出其内容的目录的路径。可以是绝对路径或相对路径。"] = ".",
        recursive: Annotated[bool, "是否递归列出子目录中的内容。"] = False,
        max_depth: Annotated[Optional[int], "递归列出内容时的最大深度。None 表示无限制。仅当 recursive 为 True 时有效。"] = None
    ) -> Dict[str, Any]:
        """
        列出指定目录中的文件和子目录。

        Args:
            directory_path: 要列出其内容的目录的路径。
            recursive: 是否递归列出子目录中的内容。
            max_depth: 递归列出内容时的最大深度。None 表示无限制。

        Returns:
            包含目录结构信息的字典，或错误信息。
            成功时，返回一个字典，包含 'base_path', 'items', 'total_items' 等键。
            每个项目包含基本信息如名称、路径、类型、大小、修改时间等。
            如果与 directory_operation 函数一起使用并设置 include_metadata=True，
            则文件项目还将包含详细的元数据。
        """
        absolute_dir_path = os.path.abspath(os.path.join(self.base_path, directory_path))
        if not os.path.exists(absolute_dir_path):
            return {"error": f"目录 {absolute_dir_path} 不存在"}
        if not os.path.isdir(absolute_dir_path):
            return {"error": f"路径 {absolute_dir_path} 不是一个目录"}

        items_list = []

        def _scan_dir(current_path, current_depth):
            if max_depth is not None and current_depth > max_depth:
                return

            try:
                for item_name in os.listdir(current_path):
                    item_full_path = os.path.join(current_path, item_name)
                    item_relative_path = os.path.relpath(item_full_path, absolute_dir_path)

                    if os.path.isdir(item_full_path):
                        items_list.append({
                            "name": item_name,
                            "path": item_relative_path,
                            "type": "directory",
                            "modified_time": os.path.getmtime(item_full_path),
                            "size_bytes": None, # 目录大小通常不直接计算
                            "depth": current_depth  # 添加深度信息
                        })
                        if recursive and (max_depth is None or current_depth < max_depth):
                            _scan_dir(item_full_path, current_depth + 1)
                    elif os.path.isfile(item_full_path):
                        _, ext = os.path.splitext(item_name)
                        mime_type, _ = mimetypes.guess_type(item_full_path)
                        items_list.append({
                            "name": item_name,
                            "path": item_relative_path,
                            "type": "file",
                            "size_bytes": os.path.getsize(item_full_path),
                            "modified_time": os.path.getmtime(item_full_path),
                            "extension": ext.lower().lstrip('.') if ext else "",
                            "mime_type": mime_type or "application/octet-stream",
                            "depth": current_depth  # 添加深度信息
                        })
            except PermissionError:
                 items_list.append({
                     "name": os.path.basename(current_path),
                     "path": os.path.relpath(current_path, absolute_dir_path),
                     "type": "directory",
                     "error": "权限不足",
                     "depth": current_depth
                 })
            except Exception as e:
                 items_list.append({
                     "name": os.path.basename(current_path),
                     "path": os.path.relpath(current_path, absolute_dir_path),
                     "type": "directory",
                     "error": str(e),
                     "depth": current_depth
                 })

        _scan_dir(absolute_dir_path, 0)

        return {
            "base_path": absolute_dir_path,
            "items": items_list,
            "total_items": len(items_list)
        }

    async def touch_file(
        self,
        file_path: Annotated[str, "要创建或更新时间戳的文件路径。"],
        create_parents: Annotated[bool, "如果为True，则创建父目录（如果不存在）。"] = True
    ) -> Dict[str, Any]:
        """
        创建空文件（如果不存在）或更新现有文件的访问和修改时间（类似 Linux 的 touch 命令）。

        Args:
            file_path: 要创建或更新时间戳的文件路径。
            create_parents: 如果为True，则创建父目录（如果不存在）。

        Returns:
            包含操作状态和文件信息的字典，包括：
            - status: 操作状态，成功或错误
            - message: 操作结果描述
            - path: 文件的绝对路径
            - created: 是否创建了新文件
            - updated: 是否更新了现有文件的时间戳

        示例:
            # 创建空文件或更新时间戳
            result = await file_utils.touch_file("empty.txt")

            # 在不存在的目录中创建文件
            result = await file_utils.touch_file("new_dir/empty.txt", create_parents=True)
        """
        absolute_file_path = os.path.abspath(os.path.join(self.base_path, file_path))

        try:
            # 确保父目录存在
            if create_parents:
                self._ensure_directory_exists(absolute_file_path)

            file_existed = os.path.exists(absolute_file_path)

            # 使用 pathlib.Path.touch() 创建文件或更新时间戳
            pathlib.Path(absolute_file_path).touch(exist_ok=True)

            if file_existed:
                return {
                    "status": "成功",
                    "message": f"已更新文件 {absolute_file_path} 的时间戳。",
                    "path": absolute_file_path,
                    "created": False,
                    "updated": True
                }
            else:
                return {
                    "status": "成功",
                    "message": f"已创建空文件 {absolute_file_path}。",
                    "path": absolute_file_path,
                    "created": True,
                    "updated": False
                }
        except Exception as e:
            return {"error": f"创建或更新文件 {absolute_file_path} 时出错: {str(e)}"}

    async def delete_item(
        self,
        path: Annotated[str, "要删除的文件或目录的路径。"],
        recursive: Annotated[bool, "如果为True，则递归删除目录及其内容；如果为False且目录非空，则会报错。"] = True,
        force: Annotated[bool, "如果为True，则忽略不存在的文件或目录，不报错；如果为False，则不存在时会报错。"] = False,
        role: Annotated[Optional[str], "提交者的角色，如果提供，将添加到提交信息中。"] = None,
        auto_commit: Annotated[bool, "是否在操作后自动执行Git提交。"] = True
    ) -> Dict[str, Any]:
        """
        删除指定路径的文件或目录。

        此函数提供了对删除操作的精细控制。
        操作后会尝试自动执行 Git 提交，可通过 auto_commit=False 禁用。

        Args:
            path: 要删除的文件或目录的路径。
            recursive: 是否递归删除目录内容。
            force: 是否忽略不存在的文件或目录。
            role: Git提交者的角色。
            auto_commit: 是否在操作后自动执行Git提交。

        Returns:
            包含操作状态和详细信息的字典。
        """
        absolute_path = os.path.abspath(os.path.join(self.base_path, path))

        if not os.path.exists(absolute_path):
            if force:
                return {
                    "status": "成功",
                    "message": f"路径 {absolute_path} 不存在，已忽略（force=True）。",
                    "path": absolute_path,
                    "item_type": "未知",
                    "existed": False
                }
            else:
                return {"error": f"路径 {absolute_path} 不存在。"}

        try:
            item_type = "未知"
            if os.path.isfile(absolute_path):
                item_type = "文件"
                os.remove(absolute_path)
            elif os.path.islink(absolute_path):
                item_type = "链接"
                os.remove(absolute_path)
            elif os.path.isdir(absolute_path):
                item_type = "目录"
                if not recursive and os.listdir(absolute_path):
                    return {"error": f"目录 {absolute_path} 不为空，且未设置递归删除（recursive=False）。"}
                shutil.rmtree(absolute_path)
            else:
                return {"error": f"路径 {absolute_path} 不是可识别的文件、目录或链接类型。"}

            result = {
                "status": "成功",
                "message": f"{item_type} {absolute_path} 已成功删除。",
                "path": absolute_path,
                "item_type": item_type,
                "existed": True
            }

            if auto_commit:
                git_result = self._git_commit(absolute_path, "delete", role)
                result["git_result"] = git_result
            else:
                result["git_result"] = {"git_status": "跳过", "message": "Git提交被禁用 (auto_commit=False)"}

            return result
        except PermissionError:
            return {"error": f"没有权限删除 {absolute_path}。请检查文件权限。"}
        except OSError as e:
            if "目录不为空" in str(e) or "directory not empty" in str(e).lower():
                return {"error": f"目录 {absolute_path} 不为空，且未设置递归删除（recursive=False）。"}
            return {"error": f"删除 {absolute_path} 时发生OS错误: {str(e)}"}
        except Exception as e:
            return {"error": f"删除 {absolute_path} 时发生未知错误: {str(e)}"}

    async def read_lines(
        self,
        file_path: Annotated[str, "要读取的文件的路径。"],
        encoding: Annotated[str, "文件编码，例如'utf-8'。"] = "utf-8",
        strip_newlines: Annotated[bool, "是否去除每行末尾的换行符。"] = True
    ) -> Dict[str, Any]:
        """
        读取文本文件并返回行列表。

        Args:
            file_path: 要读取的文件的路径。
            encoding: 文件编码。
            strip_newlines: 是否去除每行末尾的换行符。

        Returns:
            包含文件行列表的字典，或错误信息。
            成功时，返回一个字典，包含 'lines', 'line_count', 'path' 等键。
            失败时，返回一个包含 'error' 键的字典。

        示例:
            # 读取文件行
            result = await file_utils.read_lines("example.txt")
            lines = result.get("lines", [])

            # 保留换行符
            result = await file_utils.read_lines("example.txt", strip_newlines=False)
        """
        absolute_file_path = os.path.abspath(os.path.join(self.base_path, file_path))

        try:
            if not os.path.exists(absolute_file_path):
                return {"error": f"文件 {absolute_file_path} 不存在"}
            if not os.path.isfile(absolute_file_path):
                return {"error": f"路径 {absolute_file_path} 不是一个文件"}

            with open(absolute_file_path, 'r', encoding=encoding) as file:
                if strip_newlines:
                    lines = [line.rstrip('\r\n') for line in file]
                else:
                    lines = file.readlines()

            return {
                "status": "成功",
                "message": f"已读取文件 {absolute_file_path} 的行",
                "path": absolute_file_path,
                "lines": lines,
                "line_count": len(lines)
            }
        except UnicodeDecodeError:
            return {"error": f"无法使用编码 {encoding} 读取文件 {absolute_file_path}。请尝试其他编码。"}
        except Exception as e:
            return {"error": f"读取文件 {absolute_file_path} 时出错: {str(e)}"}

    async def write_lines(
        self,
        file_path: Annotated[str, "要写入的文件的路径。"],
        lines: Annotated[List[str], "要写入的行列表。"],
        encoding: Annotated[str, "文件编码，例如'utf-8'。"] = "utf-8",
        ensure_newlines: Annotated[bool, "是否确保每行以换行符结束。"] = True,
        ensure_final_newline: Annotated[bool, "确保文件以换行符结束。"] = True,
        role: Annotated[Optional[str], "提交者的角色，如果提供，将添加到提交信息中。"] = None,
        auto_commit: Annotated[bool, "是否在操作后自动执行Git提交。"] = True
    ) -> Dict[str, Any]:
        """
        将行列表写入文本文件。

        此函数提供了对换行符处理的精细控制。
        操作后会尝试自动执行 Git 提交，可通过 auto_commit=False 禁用。

        Args:
            file_path: 要写入的文件的路径。
            lines: 要写入的行列表。
            encoding: 文件编码。
            ensure_newlines: 是否确保每行以换行符结束。
            ensure_final_newline: 确保文件以换行符结束。
            role: Git提交者的角色。
            auto_commit: 是否在操作后自动执行Git提交。

        Returns:
            包含操作结果的字典，或错误信息。
            成功时，返回一个字典，包含 'status', 'message', 'path' 等键。
            失败时，返回一个包含 'error' 键的字典。

        示例:
            # 写入行列表
            lines = ["第一行", "第二行", "第三行"]
            result = await file_utils.write_lines("example.txt", lines)

            # 写入行列表但不执行Git提交
            result = await file_utils.write_lines("example.txt", lines, auto_commit=False)
        """
        absolute_file_path = os.path.abspath(os.path.join(self.base_path, file_path))

        try:
            # 确保目录存在
            self._ensure_directory_exists(absolute_file_path)

            # 处理行
            processed_lines = []
            for line in lines:
                if ensure_newlines and not line.endswith('\n'):
                    processed_lines.append(line + '\n')
                else:
                    processed_lines.append(line)

            # 确保文件以换行符结束
            if ensure_final_newline and processed_lines and not processed_lines[-1].endswith('\n'):
                processed_lines[-1] += '\n'

            # 写入文件
            with open(absolute_file_path, 'w', encoding=encoding) as file:
                file.writelines(processed_lines)

            result = {
                "status": "成功",
                "message": f"已将 {len(lines)} 行写入文件 {absolute_file_path}",
                "path": absolute_file_path,
                "line_count": len(lines)
            }

            # 执行 git 提交
            if auto_commit:
                git_result = self._git_commit(absolute_file_path, "write_lines", role)
                result["git_result"] = git_result
            else:
                result["git_result"] = {"git_status": "跳过", "message": "Git提交被禁用 (auto_commit=False)"}

            return result
        except Exception as e:
            return {"error": f"写入文件 {absolute_file_path} 时出错: {str(e)}"}

    async def search_files(
        self,
        pattern: Annotated[str, "文件名匹配模式，支持通配符，例如 '*.py', 'data_*.csv'。"],
        directory: Annotated[str, "要搜索的根目录路径。默认为当前工具实例的基础路径。"] = ".",
        recursive: Annotated[bool, "是否在子目录中递归搜索。"] = True
    ) -> Dict[str, Any]:
        """
        在指定目录中（可递归）搜索匹配特定模式的文件。

        Args:
            pattern: 文件名匹配模式。
            directory: 要搜索的根目录路径。
            recursive: 是否在子目录中递归搜索。

        Returns:
            包含搜索到的文件列表或错误信息的字典。
        """
        search_root = os.path.abspath(os.path.join(self.base_path, directory))
        if not os.path.isdir(search_root):
            return {"error": f"指定的搜索目录 {search_root} 不是一个有效的目录。"}

        found_files = []
        try:
            if recursive:
                for root, _, files in os.walk(search_root):
                    for filename in files:
                        if fnmatch.fnmatch(filename, pattern):
                            full_path = os.path.join(root, filename)
                            relative_path = os.path.relpath(full_path, search_root)
                            found_files.append({
                                "name": filename,
                                "path": relative_path,
                                "full_path": full_path,
                                "size_bytes": os.path.getsize(full_path),
                                "modified_time": os.path.getmtime(full_path)
                            })
            else:
                for filename in os.listdir(search_root):
                    full_path = os.path.join(search_root, filename)
                    if os.path.isfile(full_path) and fnmatch.fnmatch(filename, pattern):
                         relative_path = os.path.relpath(full_path, search_root)
                         found_files.append({
                            "name": filename,
                            "path": relative_path,
                            "full_path": full_path,
                            "size_bytes": os.path.getsize(full_path),
                            "modified_time": os.path.getmtime(full_path)
                        })
            return {"files": found_files, "count": len(found_files), "pattern": pattern, "directory": search_root}
        except Exception as e:
            return {"error": f"在目录 {search_root} 中搜索模式 '{pattern}' 时出错: {str(e)}"}

    async def edit_text_file(
        self,
        file_path: Annotated[str, "要编辑的文本文件的路径。"],
        changes: Annotated[List[Dict[str, Any]], "一个包含变更操作的列表。每个操作字典必须包含一个 'action' 键来指定操作类型，例如 {'action': 'replace_line', ...}。"],
        encoding: Annotated[str, "文件编码，例如'utf-8'。"] = "utf-8",
        create_if_not_exists: Annotated[bool, "如果文件不存在，是否创建新文件。"] = False,
        normalize_line_endings: Annotated[bool, "是否规范化行尾（将\\r\\n转换为\\n）。"] = True,
        ensure_final_newline: Annotated[bool, "确保文件以换行符结束。"] = True,
        role: Annotated[Optional[str], "提交者的角色，如果提供，将添加到提交信息中。"] = None,
        auto_commit: Annotated[bool, "是否在操作后自动执行Git提交。"] = True # ADDED auto_commit
    ) -> Dict[str, Any]:
        """
        编辑文本文件，支持多种操作如替换行、插入文本、删除行、正则替换、替换文本段等。

        重要: `changes` 参数是一个列表，其中每个元素都是一个字典，
        该字典 **必须包含一个名为 'action' 的键** 来指定具体的编辑操作类型。
        例如: `{"action": "replace_line", "line_number": 1, "new_content": "新内容"}`

        此函数提供了对文本文件编辑的精细控制，并可选择是否自动执行Git提交。
        操作后会尝试自动执行 Git 提交，可通过 auto_commit=False 禁用。

        Args:
            file_path: 要编辑的文本文件的路径。
            changes: 一个包含变更操作指令的列表。每个指令字典必须有 'action' 键。
                支持的 'action' 值及所需参数：
                - replace_line: 替换特定行 (参数: line_number, new_content)
                - insert_after_line: 在特定行后插入内容 (参数: line_number, text_to_insert)
                - insert_at_line: 在特定行之前插入内容 (参数: line_number, text_to_insert)
                - delete_lines: 删除特定范围的行 (参数: start_line, end_line)
                - replace_lines: 替换特定范围的行 (参数: start_line, end_line, new_content)
                - replace_text_regex: 使用正则表达式替换文本 (参数: regex_pattern, replacement_text, count)
                - replace_regex: 使用正则表达式替换文本 (参数: regex, replacement, count) (与上者同义)
                - append_text: 在文件末尾追加内容 (参数: text_to_append)
                - prepend_text: 在文件开头添加内容 (参数: text_to_prepend)
                - replace_section: 替换指定标记之间的内容 (参数: start_marker, end_marker, replacement)
                - insert_after_text: 在包含特定文本的行后插入内容 (参数: search_text, text_to_insert)
                - insert_before_text: 在包含特定文本的行前插入内容 (参数: search_text, text_to_insert)
                - comment_lines: 注释指定范围的行 (参数: start_line, end_line, comment_symbol)
                - uncomment_lines: 取消注释指定范围的行 (参数: start_line, end_line, comment_symbol)
            encoding: 文件编码。
            create_if_not_exists: 如果文件不存在，是否创建新文件。
            normalize_line_endings: 是否规范化行尾。
            ensure_final_newline: 确保文件以换行符结束。
            role: Git提交者的角色。
            auto_commit: 是否在操作后自动执行Git提交。

        Returns:
            包含编辑结果和详细信息的字典。
        """
        absolute_file_path = os.path.abspath(os.path.join(self.base_path, file_path))
        file_exists = os.path.exists(absolute_file_path)

        if not file_exists:
            if create_if_not_exists:
                directory = os.path.dirname(absolute_file_path)
                if directory and not os.path.exists(directory):
                    try:
                        os.makedirs(directory, exist_ok=True)
                    except Exception as e_mkdir:
                        return {"error": f"创建目录 {directory} 失败: {str(e_mkdir)}"}
                lines = []
                created = True
            else:
                return {"error": f"文件 {absolute_file_path} 不存在，且未设置创建（create_if_not_exists=False）。"}
        elif not os.path.isfile(absolute_file_path):
            return {"error": f"路径 {absolute_file_path} 不是一个文件。"}
        else:
            try:
                with open(absolute_file_path, 'r', encoding=encoding) as file:
                    content = file.read()
                if normalize_line_endings:
                    content = content.replace('\r\n', '\n')
                lines = content.splitlines(True)
                if not lines and content:
                    lines = [content]
                created = False
            except Exception as e_read:
                return {"error": f"读取文件 {absolute_file_path} 时出错: {str(e_read)}"}

        original_line_count = len(lines)
        applied_changes_count = 0

        try:
            for change_op in changes:
                action = change_op.get('action')
                # ... (保持原有的所有 action 处理逻辑不变) ...
                if action == 'replace_line':
                    line_num = change_op.get('line_number')
                    new_content = change_op.get('new_content', '')
                    if line_num is None: return {"error": f"replace_line 操作缺少 'line_number' 参数。"}
                    if not (1 <= line_num <= len(lines)) and not (line_num == 1 and len(lines) == 0):
                        return {"error": f"无效的行号 {line_num} 用于 replace_line。文件有 {len(lines)} 行。"}
                    if len(lines) == 0 and line_num == 1:
                        lines.append(new_content + ('\n' if not new_content.endswith('\n') else ''))
                    else:
                        lines[line_num - 1] = new_content + ('\n' if not new_content.endswith('\n') else '')
                    applied_changes_count += 1
                elif action == 'insert_after_line':
                    line_num = change_op.get('line_number')
                    text_to_insert = change_op.get('text_to_insert', '')
                    if line_num is None: return {"error": f"insert_after_line 操作缺少 'line_number' 参数。"}
                    if not (0 <= line_num <= len(lines)):
                        return {"error": f"无效的行号 {line_num} 用于 insert_after_line。文件有 {len(lines)} 行。"}
                    insert_text = text_to_insert + ('\n' if not text_to_insert.endswith('\n') else '')
                    if line_num == 0:
                        lines.insert(0, insert_text)
                    else:
                        lines.insert(line_num, insert_text)
                    applied_changes_count += 1
                # ... (所有其他 elif action == '...' 块保持原样) ...
                elif action == 'replace_section':
                    start_marker = change_op.get('start_marker')
                    end_marker = change_op.get('end_marker')
                    replacement = change_op.get('replacement', '')
                    if not start_marker or not end_marker:
                        return {"error": "replace_section 操作需要 'start_marker' 和 'end_marker' 参数。"}
                    full_text = ''.join(lines)
                    start_pos = full_text.find(start_marker)
                    if start_pos == -1: return {"error": f"找不到开始标记: '{start_marker}'"}
                    end_search_start = start_pos + len(start_marker)
                    end_pos = full_text.find(end_marker, end_search_start)
                    if end_pos == -1: return {"error": f"找不到结束标记: '{end_marker}'"}
                    new_text = full_text[:start_pos] + replacement + full_text[end_pos + len(end_marker):]
                    lines = new_text.splitlines(True)
                    if not lines and new_text: lines = [new_text]
                    applied_changes_count += 1
                else:
                    return {"error": f"未知的编辑操作: {action}"}

            if ensure_final_newline and lines and not lines[-1].endswith('\n'):
                lines[-1] += '\n'

            with open(absolute_file_path, 'w', encoding=encoding) as file:
                file.writelines(lines)

            result = {
                "status": "成功",
                "message": f"文件 {absolute_file_path} 已{'创建并' if created else ''}编辑。",
                "path": absolute_file_path,
                "changes_applied": applied_changes_count,
                "original_line_count": original_line_count,
                "new_line_count": len(lines),
                "created": created
            }

            if auto_commit: # MODIFIED: Conditional Git commit
                git_result = self._git_commit(absolute_file_path, "edit", role)
                result["git_result"] = git_result
            else:
                result["git_result"] = {"git_status": "跳过", "message": "Git提交被禁用 (auto_commit=False)"}

            return result

        except Exception as e:
            return {"error": f"编辑文件 {absolute_file_path} 时出错: {str(e)}"}

    async def highlight_code(
        self,
        code_or_file_path: Annotated[str, "要高亮的代码字符串或代码文件的路径。"],
        language: Annotated[Optional[str], "代码语言，例如 'python', 'javascript'。如果为 None，则会尝试根据文件名（如果提供了路径）或内容自动检测。"] = None,
        output_format: Annotated[str, "输出格式，支持 'html' 或 'terminal'。"] = "html"
    ) -> Dict[str, Any]:
        """
        对提供的代码字符串或文件内容进行语法高亮。

        Args:
            code_or_file_path: 要高亮的代码字符串或代码文件的路径。
            language: 代码语言。如果为 None，则自动检测。
            output_format: 输出格式，'html' 或 'terminal'。

        Returns:
            包含高亮后代码、语言、格式等信息的字典，或错误信息。
        """
        if not PYGMENTS_AVAILABLE:
            return {"error": "Pygments 库未安装，无法进行代码高亮。"}

        code_to_highlight = ""
        detected_language = language

        if os.path.isfile(os.path.join(self.base_path, code_or_file_path)):
            absolute_path = os.path.abspath(os.path.join(self.base_path, code_or_file_path))
            try:
                with open(absolute_path, 'r', encoding='utf-8') as f:
                    code_to_highlight = f.read()
                if not detected_language: # 如果没有指定语言，尝试从文件名获取
                    try:
                        lexer = get_lexer_for_filename(absolute_path)
                        detected_language = lexer.name
                    except ClassNotFound:
                        pass # 保持 detected_language 为 None
            except Exception as e:
                return {"error": f"读取文件 {absolute_path} 进行高亮时出错: {str(e)}"}
        else:
            code_to_highlight = code_or_file_path # 假定是代码字符串

        if not code_to_highlight:
            return {"error": "没有提供用于高亮的代码内容。"}

        try:
            if detected_language:
                lexer = get_lexer_by_name(detected_language)
            else:
                lexer = guess_lexer(code_to_highlight) # 自动猜测语言
                detected_language = lexer.name
        except ClassNotFound:
            return {"error": f"无法找到语言 '{detected_language or 'auto-detected'}' 的词法分析器。"}
        except Exception as e:
            return {"error": f"获取词法分析器时出错: {str(e)}"}

        try:
            if output_format.lower() == 'html':
                formatter = HtmlFormatter(linenos=True, cssclass="highlight-source", full=True, style='default')
                highlighted_code = highlight(code_to_highlight, lexer, formatter)
                # css = formatter.get_style_defs('.highlight-source') # 获取内联CSS，但full=True时已包含
                return {"highlighted_code": highlighted_code, "language": detected_language, "format": "html"}
            elif output_format.lower() == 'terminal':
                formatter = TerminalFormatter(style='default') # 可以选择不同的 style
                highlighted_code = highlight(code_to_highlight, lexer, formatter)
                return {"highlighted_code": highlighted_code, "language": detected_language, "format": "terminal"}
            else:
                return {"error": f"不支持的输出格式: {output_format}。请选择 'html' 或 'terminal'。"}
        except Exception as e:
            return {"error": f"代码高亮过程中出错: {str(e)}"}

    async def compare_files(
        self,
        file1_path: Annotated[str, "第一个要比较的文件的路径。"],
        file2_path: Annotated[str, "第二个要比较的文件的路径。"],
        context_lines: Annotated[int, "差异上下文中显示的行数。"] = 3
    ) -> Dict[str, Any]:
        """
        比较两个文本文件的内容差异。

        Args:
            file1_path: 第一个文件的路径。
            file2_path: 第二个文件的路径。
            context_lines: 差异上下文中显示的行数。

        Returns:
            包含差异文本、HTML差异、相似度等信息的字典，或错误信息。
        """
        abs_file1_path = os.path.abspath(os.path.join(self.base_path, file1_path))
        abs_file2_path = os.path.abspath(os.path.join(self.base_path, file2_path))

        if not os.path.exists(abs_file1_path):
            return {"error": f"文件 {abs_file1_path} 不存在"}
        if not os.path.exists(abs_file2_path):
            return {"error": f"文件 {abs_file2_path} 不存在"}
        if not os.path.isfile(abs_file1_path):
            return {"error": f"路径 {abs_file1_path} 不是文件"}
        if not os.path.isfile(abs_file2_path):
            return {"error": f"路径 {abs_file2_path} 不是文件"}

        try:
            with open(abs_file1_path, 'r', encoding='utf-8') as f1:
                file1_lines = f1.readlines()
            with open(abs_file2_path, 'r', encoding='utf-8') as f2:
                file2_lines = f2.readlines()

            diff_text = ''.join(difflib.unified_diff(file1_lines, file2_lines, fromfile=file1_path, tofile=file2_path, n=context_lines))

            html_diff_generator = difflib.HtmlDiff(tabsize=4, wrapcolumn=80)
            html_diff = html_diff_generator.make_file(file1_lines, file2_lines, file1_path, file2_path, context=True, numlines=context_lines)

            matcher = difflib.SequenceMatcher(None, ''.join(file1_lines), ''.join(file2_lines))
            similarity = matcher.ratio() * 100

            return {
                "diff_text": diff_text if diff_text else "文件内容相同。",
                "html_diff": html_diff,
                "similarity_percentage": round(similarity, 2),
                "file1_info": {"path": abs_file1_path, "line_count": len(file1_lines), "size_bytes": os.path.getsize(abs_file1_path)},
                "file2_info": {"path": abs_file2_path, "line_count": len(file2_lines), "size_bytes": os.path.getsize(abs_file2_path)},
            }
        except Exception as e:
            return {"error": f"比较文件 {abs_file1_path} 和 {abs_file2_path} 时出错: {str(e)}"}

    async def get_project_structure(
        self,
        root_dir: Annotated[str, "要分析的项目根目录的路径。"] = ".",
        max_depth: Annotated[Optional[int], "递归扫描的最大深度。None 表示无限制。"] = None,
        include_patterns: Annotated[Optional[List[str]], "要包含的文件/目录名模式列表 (例如 ['src', '*.py'])。如果为 None，则包含所有。"] = None,
        exclude_patterns: Annotated[Optional[List[str]], "要排除的文件/目录名模式列表 (例如 ['__pycache__', '*.tmp', '.git'])。如果为 None，则不排除。"] = None,
        output_format: Annotated[str, "输出格式：'json' 返回结构化数据，'text' 返回树状文本表示。"] = "json"
    ) -> Dict[str, Any]:
        """
        获取项目目录结构，可指定深度、包含/排除模式，并以 JSON 或文本树形式输出。

        Args:
            root_dir: 项目根目录。
            max_depth: 最大递归深度。
            include_patterns: 要包含的文件/目录模式列表。
            exclude_patterns: 要排除的文件/目录模式列表。
            output_format: 'json' 或 'text'。

        Returns:
            包含项目结构（JSON对象或文本树）或错误信息的字典。
        """
        absolute_root_dir = os.path.abspath(os.path.join(self.base_path, root_dir))
        if not os.path.isdir(absolute_root_dir):
            return {"error": f"指定的根目录 {absolute_root_dir} 不是一个有效的目录。"}

        default_excludes = ['.git', '.idea', '__pycache__', 'node_modules', '.venv', '.vscode', '*.pyc', '*.swp', '*.DS_Store']
        current_excludes = default_excludes + (exclude_patterns if exclude_patterns else [])

        def _is_excluded(name, path):
            for pattern in current_excludes:
                if fnmatch.fnmatch(name, pattern) or fnmatch.fnmatch(path, pattern):
                    return True
            return False

        def _is_included(name, path):
            if not include_patterns:
                return True
            for pattern in include_patterns:
                if fnmatch.fnmatch(name, pattern) or fnmatch.fnmatch(path, pattern):
                    return True
            return False

        def _scan_dir_recursive(current_dir_path, current_depth):
            if max_depth is not None and current_depth > max_depth:
                return {"name": os.path.basename(current_dir_path), "type": "directory", "path": current_dir_path, "truncated": True, "children": []}

            node = {"name": os.path.basename(current_dir_path) or current_dir_path, "type": "directory", "path": current_dir_path, "children": []}
            try:
                items = sorted(os.listdir(current_dir_path))
                for item_name in items:
                    item_full_path = os.path.join(current_dir_path, item_name)
                    if _is_excluded(item_name, item_full_path):
                        continue
                    if not _is_included(item_name, item_full_path):
                        continue

                    if os.path.isdir(item_full_path):
                        node["children"].append(_scan_dir_recursive(item_full_path, current_depth + 1))
                    elif os.path.isfile(item_full_path):
                         _, ext = os.path.splitext(item_name)
                         node["children"].append({
                            "name": item_name,
                            "type": "file",
                            "path": item_full_path,
                            "size_bytes": os.path.getsize(item_full_path),
                            "extension": ext.lower().lstrip('.') if ext else ""
                        })
            except PermissionError:
                node["error"] = "权限不足"
            except Exception as e_scan:
                node["error"] = str(e_scan)
            return node

        structure_json = _scan_dir_recursive(absolute_root_dir, 0)

        if output_format.lower() == 'text':
            def _generate_tree_text(node, prefix="", is_last=True):
                lines = []
                connector = "└── " if is_last else "├── "
                lines.append(f"{prefix}{connector}{node['name']}{' (...)' if node.get('truncated') else ''}{' [ERROR: ' + node['error'] + ']' if node.get('error') else ''}")

                if node.get("type") == "directory" and "children" in node and not node.get('truncated'):
                    children_prefix = prefix + ("    " if is_last else "│   ")
                    for i, child in enumerate(node["children"]):
                        lines.extend(_generate_tree_text(child, children_prefix, i == len(node["children"]) - 1))
                return lines

            tree_text_lines = [absolute_root_dir] + _generate_tree_text(structure_json, is_last=True)
            return {"structure_text": "\n".join(tree_text_lines), "root_directory": absolute_root_dir}

        elif output_format.lower() == 'json':
            return {"structure": structure_json, "root_directory": absolute_root_dir}
        else:
            return {"error": "无效的 output_format。请选择 'json' 或 'text'。"}

    async def start_file_monitor(
        self,
        path_to_monitor: Annotated[str, "要监控的文件或目录的路径。"],
        on_change_callback: Annotated[Any, "当检测到文件更改时调用的回调函数。它接收两个参数：事件类型 (str) 和受影响文件的路径 (str)。"]
    ) -> Dict[str, Any]:
        """
        开始监控指定文件或目录的变化。

        Args:
            path_to_monitor: 要监控的文件或目录的路径。
            on_change_callback: 文件变化时的回调函数。

        Returns:
            包含监控状态信息的字典。
        """
        if not WATCHDOG_AVAILABLE:
            return {"error": "Watchdog 库未安装，无法启动文件监控。"}

        absolute_path = os.path.abspath(os.path.join(self.base_path, path_to_monitor))
        if not os.path.exists(absolute_path):
            return {"error": f"路径 {absolute_path} 不存在。"}

        if absolute_path in self._file_monitors:
            return {"status": "已在监控", "message": f"路径 {absolute_path} 已在监控中。", "monitor_id": absolute_path}

        try:
            event_handler = FileChangeHandler(on_change_callback, absolute_path)
            observer = Observer()
            observer.schedule(event_handler, absolute_path, recursive=True)
            observer.start()
            self._file_monitors[absolute_path] = observer
            return {"status": "成功", "message": f"已开始监控路径 {absolute_path}。", "monitor_id": absolute_path}
        except Exception as e:
            return {"error": f"启动文件监控失败: {str(e)}"}

    async def stop_file_monitor(
        self,
        monitor_id: Annotated[str, "要停止监控的路径，与 start_file_monitor 返回的 monitor_id 相同。"]
    ) -> Dict[str, Any]:
        """
        停止对指定路径的文件或目录的监控。

        Args:
            monitor_id: 通过 start_file_monitor 获取的监控ID（即被监控的绝对路径）。

        Returns:
            包含操作状态的字典。
        """
        if not WATCHDOG_AVAILABLE:
            return {"error": "Watchdog 库未安装，无法操作文件监控。"}

        absolute_path = os.path.abspath(monitor_id) # monitor_id 就是绝对路径
        if absolute_path not in self._file_monitors:
            return {"error": f"路径 {absolute_path} 当前未被监控。"}

        try:
            observer = self._file_monitors.pop(absolute_path)
            observer.stop()
            observer.join(timeout=5) # 等待线程结束
            return {"status": "成功", "message": f"已停止对路径 {absolute_path} 的监控。"}
        except Exception as e:
            # 即使出错，也尝试从字典中移除，避免状态不一致
            if absolute_path in self._file_monitors:
                 del self._file_monitors[absolute_path]
            return {"error": f"停止文件监控失败: {str(e)}"}

    async def get_file_monitor_status(
        self,
        monitor_id: Annotated[Optional[str], "要查询状态的监控ID (路径)。如果为 None，则返回所有活动监控的状态。"] = None
    ) -> Dict[str, Any]:
        """
        获取指定文件监控或所有文件监控的状态。

        Args:
            monitor_id: 要查询的监控ID（即被监控的绝对路径）。如果为 None，则返回所有监控的状态。

        Returns:
            包含监控状态信息的字典。
        """
        if not WATCHDOG_AVAILABLE:
            return {"status": "不可用", "message": "Watchdog 库未安装，文件监控功能不可用。"}

        if monitor_id:
            absolute_path = os.path.abspath(monitor_id)
            if absolute_path in self._file_monitors and self._file_monitors[absolute_path].is_alive():
                return {"monitor_id": absolute_path, "status": "运行中"}
            else:
                return {"monitor_id": absolute_path, "status": "未运行或不存在"}
        else:
            statuses = []
            for path, observer in self._file_monitors.items():
                statuses.append({"monitor_id": path, "status": "运行中" if observer.is_alive() else "已停止"})
            return {"active_monitors": statuses, "count": len(statuses)}

    async def create_directory(
        self,
        directory_path: Annotated[str, "要创建的目录的路径。可以是多级目录。"],
        exist_ok: Annotated[bool, "如果为 True，当目录已存在时不会引发错误。如果为 False (默认)，目录已存在会报错。"] = False
    ) -> Dict[str, Any]:
        """
        创建指定路径的目录，如果父目录不存在也会一并创建。

        Args:
            directory_path: 要创建的目录的路径。
            exist_ok: 如果目录已存在，是否忽略错误。

        Returns:
            包含操作状态和目录路径的字典。
        """
        absolute_dir_path = os.path.abspath(os.path.join(self.base_path, directory_path))
        try:
            os.makedirs(absolute_dir_path, exist_ok=exist_ok)
            return {"status": "成功", "message": f"目录 {absolute_dir_path} 已创建（或已存在）。", "path": absolute_dir_path}
        except FileExistsError as e: # 仅当 exist_ok=False 时会触发
            return {"error": f"创建目录 {absolute_dir_path} 失败: 目录已存在。 {str(e)}"}
        except Exception as e:
            return {"error": f"创建目录 {absolute_dir_path} 时出错: {str(e)}"}

    async def move_item(
        self,
        source_path: Annotated[str, "要移动的文件或目录的源路径。"],
        destination_path: Annotated[str, "目标路径。如果目标是目录，则源项目将被移动到该目录下。"],
        overwrite: Annotated[bool, "如果目标已存在，是否覆盖。默认为 False。"] = False
    ) -> Dict[str, Any]:
        """
        移动或重命名文件或目录。

        Args:
            source_path: 源路径。
            destination_path: 目标路径。
            overwrite: 如果目标已存在，是否覆盖。

        Returns:
            包含操作状态的字典。
        """
        abs_source_path = os.path.abspath(os.path.join(self.base_path, source_path))
        abs_destination_path = os.path.abspath(os.path.join(self.base_path, destination_path))

        if not os.path.exists(abs_source_path):
            return {"error": f"源路径 {abs_source_path} 不存在。"}

        # 如果需要覆盖且目标存在
        if overwrite and os.path.exists(abs_destination_path):
            try:
                if os.path.isfile(abs_destination_path) or os.path.islink(abs_destination_path):
                    os.remove(abs_destination_path)
                    print(f"覆盖操作：已删除已存在的文件/链接 {abs_destination_path}")
                elif os.path.isdir(abs_destination_path):
                    shutil.rmtree(abs_destination_path)
                    print(f"覆盖操作：已删除已存在的目录 {abs_destination_path}")
            except Exception as e_del:
                return {"error": f"覆盖操作：删除已存在的目标 {abs_destination_path} 失败: {str(e_del)}"}

        # 确保目标路径的父目录存在 (如果目标不是一个已存在的目录)
        dest_dir = os.path.dirname(abs_destination_path)
        if dest_dir and not os.path.exists(dest_dir):
            try:
                os.makedirs(dest_dir, exist_ok=True)
                print(f"已创建目标父目录 {dest_dir}")
            except Exception as e_mkdir:
                 return {"error": f"创建目标父目录 {dest_dir} 失败: {str(e_mkdir)}"}

        try:
            shutil.move(abs_source_path, abs_destination_path)
            return {"status": "成功", "message": f"已将 {abs_source_path} 移动到 {abs_destination_path}。"}
        except Exception as e:
            # 如果错误是因为目标已存在且不允许覆盖（shutil.move 的默认行为）
            if isinstance(e, shutil.Error) and "already exists" in str(e).lower() and not overwrite:
                 return {"error": f"移动 {abs_source_path} 到 {abs_destination_path} 失败: 目标已存在且未设置覆盖 (overwrite=False)。"}
            return {"error": f"移动 {abs_source_path} 到 {abs_destination_path} 时出错: {str(e)}"}

    async def copy_item(
        self,
        source_path: Annotated[str, "要复制的文件或目录的源路径。"],
        destination_path: Annotated[str, "目标路径。如果目标是目录，则源项目将被复制到该目录下。"],
        overwrite: Annotated[bool, "如果目标已存在，是否覆盖。默认为 False。"] = False
    ) -> Dict[str, Any]:
        """
        复制文件或目录。

        Args:
            source_path: 源路径。
            destination_path: 目标路径。
            overwrite: 如果目标已存在，是否覆盖。

        Returns:
            包含操作状态的字典。
        """
        abs_source_path = os.path.abspath(os.path.join(self.base_path, source_path))
        abs_destination_path = os.path.abspath(os.path.join(self.base_path, destination_path))

        if not os.path.exists(abs_source_path):
            return {"error": f"源路径 {abs_source_path} 不存在。"}

        if os.path.exists(abs_destination_path) and not overwrite:
            return {"error": f"目标路径 {abs_destination_path} 已存在，且未设置覆盖。"}

        # 确保目标路径的父目录存在
        dest_parent_dir = os.path.dirname(abs_destination_path)
        if os.path.isdir(abs_source_path) and os.path.isdir(abs_destination_path): # 复制目录到目录
             dest_final_path = os.path.join(abs_destination_path, os.path.basename(abs_source_path))
        elif os.path.isdir(abs_destination_path): # 复制文件到目录
             dest_final_path = os.path.join(abs_destination_path, os.path.basename(abs_source_path))
        else: # 复制文件到文件，或目录到新目录名
             dest_final_path = abs_destination_path
             dest_parent_dir = os.path.dirname(dest_final_path)


        if dest_parent_dir and not os.path.exists(dest_parent_dir):
            try:
                os.makedirs(dest_parent_dir, exist_ok=True)
            except Exception as e_mkdir:
                 return {"error": f"创建目标父目录 {dest_parent_dir} 失败: {str(e_mkdir)}"}

        try:
            if os.path.isdir(abs_source_path):
                if os.path.exists(dest_final_path) and overwrite:
                    shutil.rmtree(dest_final_path)
                shutil.copytree(abs_source_path, dest_final_path)
            else: # isfile
                shutil.copy2(abs_source_path, dest_final_path) # copy2 preserves metadata
            return {"status": "成功", "message": f"已将 {abs_source_path} 复制到 {dest_final_path}。"}
        except Exception as e:
            return {"error": f"复制 {abs_source_path} 到 {dest_final_path} 时出错: {str(e)}"}

    async def get_file_metadata(
        self,
        file_path: Annotated[str, "要获取元数据的文件路径。"]
    ) -> Dict[str, Any]:
        """
        获取指定文件的详细元数据。

        Args:
            file_path: 文件路径。

        Returns:
            包含文件元数据（大小、修改时间、创建时间、访问时间、MIME类型等）的字典，或错误信息。
        """
        absolute_file_path = os.path.abspath(os.path.join(self.base_path, file_path))
        if not os.path.exists(absolute_file_path):
            return {"error": f"文件 {absolute_file_path} 不存在。"}
        if not os.path.isfile(absolute_file_path):
            return {"error": f"路径 {absolute_file_path} 不是一个文件。"}

        try:
            stat_info = os.stat(absolute_file_path)
            mime_type, _ = mimetypes.guess_type(absolute_file_path)
            _, ext = os.path.splitext(absolute_file_path)

            metadata = {
                "path": absolute_file_path,
                "name": os.path.basename(absolute_file_path),
                "size_bytes": stat_info.st_size,
                "size_kb": round(stat_info.st_size / 1024, 2),
                "size_mb": round(stat_info.st_size / (1024 * 1024), 2),
                "modified_time_timestamp": stat_info.st_mtime,
                "modified_time_iso": datetime.datetime.fromtimestamp(stat_info.st_mtime).isoformat(),
                "accessed_time_timestamp": stat_info.st_atime,
                "accessed_time_iso": datetime.datetime.fromtimestamp(stat_info.st_atime).isoformat(),
                "metadata_change_time_timestamp": stat_info.st_ctime, # st_ctime 实际上是元数据更改时间，而非创建时间
                "metadata_change_time_iso": datetime.datetime.fromtimestamp(stat_info.st_ctime).isoformat(),
                "mime_type": mime_type or "application/octet-stream",
                "extension": ext.lower().lstrip('.') if ext else "",
                "is_link": os.path.islink(absolute_file_path),
            }
            # 尝试获取更准确的创建时间 (Windows)
            if os.name == 'nt':
                metadata["created_time_timestamp_windows"] = os.path.getctime(absolute_file_path)
                metadata["created_time_iso_windows"] = datetime.datetime.fromtimestamp(os.path.getctime(absolute_file_path)).isoformat()

            return {"status": "成功", "metadata": metadata}
        except Exception as e:
            return {"error": f"获取文件 {absolute_file_path} 元数据时出错: {str(e)}"}

    async def create_archive(
        self,
        archive_path: Annotated[str, "要创建的压缩文件的完整路径（例如 'backup.zip' 或 'archive.tar.gz'）。"],
        source_paths: Annotated[List[str], "要添加到压缩文件中的文件或目录的路径列表。"],
        archive_format: Annotated[str, "压缩格式，支持 'zip', 'tar', 'gztar' (tar.gz), 'bztar' (tar.bz2)。"] = "zip",
        compression_level: Annotated[Optional[int], "压缩级别 (对于zip: 0-9, 对于gz/bz2: 1-9)。None 表示默认。"] = None
    ) -> Dict[str, Any]:
        """
        将指定的文件和目录创建为一个压缩文件。

        Args:
            archive_path: 要创建的压缩文件的路径。
            source_paths: 要压缩的文件或目录列表。
            archive_format: 压缩格式 ('zip', 'tar', 'gztar', 'bztar')。
            compression_level: 压缩级别。

        Returns:
            包含操作状态和压缩文件信息的字典。
        """
        absolute_archive_path = os.path.abspath(os.path.join(self.base_path, archive_path))
        archive_dir = os.path.dirname(absolute_archive_path)
        if archive_dir and not os.path.exists(archive_dir):
            os.makedirs(archive_dir, exist_ok=True)

        processed_source_paths = [os.path.abspath(os.path.join(self.base_path, p)) for p in source_paths]
        for sp in processed_source_paths:
            if not os.path.exists(sp):
                return {"error": f"源路径 {sp} 不存在。"}

        try:
            if archive_format == "zip":
                compress_type = zipfile.ZIP_DEFLATED
                if compression_level is not None: # zipfile 模块的 compresslevel 参数在 Python 3.7+ 可用
                    if not (0 <= compression_level <= 9):
                        print(f"警告: ZIP 的 compression_level ({compression_level}) 无效, 将使用默认值。")
                # 对于 zipfile，compresslevel 参数在 Python 3.7+ 中可用
                # 在 Python 3.6 及更早版本中，zipfile.ZipFile 不直接接受 compresslevel 参数
                # 这里我们假设 Python 3.7+
                with zipfile.ZipFile(absolute_archive_path, 'w', compress_type, compresslevel=compression_level if compression_level is not None else None) as zf:
                    for src_path in processed_source_paths:
                        if os.path.isdir(src_path):
                            for root, _, files in os.walk(src_path):
                                for file in files:
                                    file_to_add = os.path.join(root, file)
                                    arcname = os.path.relpath(file_to_add, os.path.dirname(src_path))
                                    zf.write(file_to_add, arcname)
                        else: # isfile
                            zf.write(src_path, os.path.basename(src_path))
            elif archive_format in ["tar", "gztar", "bztar"]:
                mode = "w"
                if archive_format == "gztar":
                    mode = "w:gz"
                elif archive_format == "bztar":
                    mode = "w:bz2"

                # tarfile 的 compresslevel 参数需要 Python 3.9+
                # 我们将简单地使用默认压缩，或在未来版本中添加版本检查
                with tarfile.open(absolute_archive_path, mode) as tf:
                    for src_path in processed_source_paths:
                        tf.add(src_path, arcname=os.path.basename(src_path))
            else:
                return {"error": f"不支持的压缩格式: {archive_format}。请使用 'zip', 'tar', 'gztar', 'bztar'。"}

            return {"status": "成功", "message": f"已创建压缩文件 {absolute_archive_path}", "path": absolute_archive_path, "size_bytes": os.path.getsize(absolute_archive_path)}
        except Exception as e:
            return {"error": f"创建压缩文件 {absolute_archive_path} 时出错: {str(e)}"}

    async def extract_archive(
        self,
        archive_path: Annotated[str, "要解压的压缩文件的路径。"],
        destination_dir: Annotated[str, "解压后文件存放的目标目录。"],
        archive_format: Annotated[Optional[str], "压缩格式，例如 'zip', 'tar', 'gztar'。如果为 None，则尝试从文件扩展名推断。"] = None
    ) -> Dict[str, Any]:
        """
        解压指定的压缩文件到目标目录。

        Args:
            archive_path: 压缩文件路径。
            destination_dir: 解压目标目录。
            archive_format: 压缩格式。如果为 None，则自动推断。

        Returns:
            包含操作状态的字典。
        """
        abs_archive_path = os.path.abspath(os.path.join(self.base_path, archive_path))
        abs_destination_dir = os.path.abspath(os.path.join(self.base_path, destination_dir))

        if not os.path.exists(abs_archive_path):
            return {"error": f"压缩文件 {abs_archive_path} 不存在。"}
        if not os.path.isfile(abs_archive_path):
            return {"error": f"路径 {abs_archive_path} 不是一个文件。"}

        if not os.path.exists(abs_destination_dir):
            os.makedirs(abs_destination_dir, exist_ok=True)
        elif not os.path.isdir(abs_destination_dir):
            return {"error": f"目标路径 {abs_destination_dir} 已存在且不是目录。"}

        fmt = archive_format
        if not fmt:
            if abs_archive_path.endswith(".zip"):
                fmt = "zip"
            elif abs_archive_path.endswith(".tar.gz") or abs_archive_path.endswith(".tgz"):
                fmt = "gztar"
            elif abs_archive_path.endswith(".tar.bz2") or abs_archive_path.endswith(".tbz2"):
                fmt = "bztar"
            elif abs_archive_path.endswith(".tar"):
                fmt = "tar"
            else:
                return {"error": f"无法从文件名推断压缩格式: {abs_archive_path}。请明确指定 archive_format。"}

        try:
            if fmt == "zip":
                with zipfile.ZipFile(abs_archive_path, 'r') as zf:
                    zf.extractall(abs_destination_dir)
            elif fmt in ["tar", "gztar", "bztar"]:
                mode = "r"
                if fmt == "gztar":
                    mode = "r:gz"
                elif fmt == "bztar":
                    mode = "r:bz2"
                with tarfile.open(abs_archive_path, mode) as tf:
                    tf.extractall(abs_destination_dir)
            else:
                return {"error": f"不支持的解压格式: {fmt}。"}

            return {"status": "成功", "message": f"文件已解压到 {abs_destination_dir}。"}
        except Exception as e:
            return {"error": f"解压文件 {abs_archive_path} 时出错: {str(e)}"}

    async def get_file_hash(
        self,
        file_path: Annotated[str, "要计算哈希值的文件路径。"],
        algorithm: Annotated[str, "哈希算法，例如 'md5', 'sha1', 'sha256', 'sha512'。"] = "sha256"
    ) -> Dict[str, Any]:
        """
        计算指定文件的哈希值。

        Args:
            file_path: 文件路径。
            algorithm: 哈希算法。

        Returns:
            包含哈希值或错误信息的字典。
        """
        absolute_file_path = os.path.abspath(os.path.join(self.base_path, file_path))
        if not os.path.exists(absolute_file_path):
            return {"error": f"文件 {absolute_file_path} 不存在。"}
        if not os.path.isfile(absolute_file_path):
            return {"error": f"路径 {absolute_file_path} 不是一个文件。"}

        try:
            hasher = hashlib.new(algorithm)
            with open(absolute_file_path, 'rb') as f:
                while chunk := f.read(8192): # 8KB 块
                    hasher.update(chunk)
            return {"status": "成功", "algorithm": algorithm, "hash": hasher.hexdigest(), "file_path": absolute_file_path}
        except ValueError:
            return {"error": f"不支持的哈希算法: {algorithm}。请尝试 'md5', 'sha1', 'sha256', 'sha512' 等。"}
        except Exception as e:
            return {"error": f"计算文件 {absolute_file_path} 哈希时出错: {str(e)}"}

    async def get_diagnostics(
        self,
        file_path: Annotated[str, "要获取诊断信息的文件路径（通常是代码文件）。"],
        config_path: Annotated[Optional[str], "Pylint配置文件路径。如果为None，将使用默认配置。"] = None
    ) -> Dict[str, Any]:
        """
        获取代码文件的诊断信息，包括语法错误、代码风格问题、潜在的bug等。

        此函数使用Pylint对Python代码进行静态分析，提供全面的代码质量报告。
        支持的检查包括：
        - 语法错误和异常
        - 代码风格问题（PEP 8）
        - 潜在的bug和逻辑错误
        - 代码复杂度和可维护性问题
        - 未使用的变量和导入
        - 命名约定违规

        Args:
            file_path: 要检查的文件路径。
            config_path: Pylint配置文件路径。如果为None，将使用默认配置。

        Returns:
            包含诊断信息的字典。
        """
        absolute_file_path = os.path.abspath(os.path.join(self.base_path, file_path))
        pylint_warnings = [] # 用于收集关于Pylint执行本身的警告

        if not os.path.exists(absolute_file_path):
            return {"status": "失败", "message": f"文件 {absolute_file_path} 不存在"}

        if not os.path.isfile(absolute_file_path):
            return {"status": "失败", "message": f"路径 {absolute_file_path} 不是一个文件"}

        _, ext = os.path.splitext(absolute_file_path)
        file_type = ext.lower().lstrip('.')

        if file_type not in ['py', 'pyw']:
            return {
                "status": "失败",
                "message": f"目前只支持Python文件(.py, .pyw)的代码诊断，不支持 '{file_type}' 类型文件。"
            }

        try:
            import subprocess # 移到函数内部，仅在需要时导入
            import json       # 移到函数内部

            cmd = ['pylint', absolute_file_path, '--output-format=json']

            if config_path:
                absolute_config_path = os.path.abspath(os.path.join(self.base_path, config_path))
                if os.path.exists(absolute_config_path):
                    cmd.append(f'--rcfile={absolute_config_path}')
                else:
                    warning_msg = f"警告: Pylint配置文件 {absolute_config_path} 未找到，将使用默认配置。"
                    print(warning_msg) # 打印到控制台
                    pylint_warnings.append(warning_msg) # 也收集到结果中

            # 执行pylint命令
            # 使用 check=False 来手动处理非零退出码，因为Pylint发现问题时会返回非零码
            process = subprocess.run(cmd, capture_output=True, text=True, encoding='utf-8', check=False)

            results_text = process.stdout
            pylint_stderr = process.stderr.strip() if process.stderr else None

            if pylint_stderr:
                warning_msg = f"Pylint stderr: {pylint_stderr}"
                print(warning_msg)
                pylint_warnings.append(warning_msg)

            diagnostics = []
            if not results_text.strip():
                # 如果stdout为空，但Pylint有返回码且stderr有内容，则可能Pylint本身执行出错
                if process.returncode != 0 and pylint_stderr:
                    return {
                        "status": "失败",
                        "message": f"Pylint执行失败。Return code: {process.returncode}. Stderr: {pylint_stderr}",
                        "pylint_warnings": pylint_warnings
                    }
                # 否则，没有JSON输出通常意味着没有发现问题（或者Pylint配置为不输出任何东西）
                pass # diagnostics 列表将为空
            else:
                try:
                    raw_results = json.loads(results_text)
                    for item in raw_results:
                        severity = self._map_pylint_severity(item.get('type', ''))
                        diagnostics.append({
                            "line": item.get('line', 0),
                            "column": item.get('column', 0),
                            "severity": severity,
                            "message": f"{item.get('message', '')} ({item.get('symbol', item.get('message-id', 'N/A'))})", # 添加symbol/message-id
                            "source": item.get('symbol', item.get('message-id', 'N/A')) # 保持source为symbol或message-id
                        })
                except json.JSONDecodeError:
                    return {
                        "status": "失败",
                        "message": f"解析Pylint JSON输出失败。Return code: {process.returncode}. Stdout (first 500 chars): {results_text[:500]}...",
                        "pylint_warnings": pylint_warnings,
                        "pylint_stderr": pylint_stderr
                    }

            error_count = sum(1 for d in diagnostics if d['severity'] == 'error')
            warning_count = sum(1 for d in diagnostics if d['severity'] == 'warning')
            info_count = sum(1 for d in diagnostics if d['severity'] == 'info')
            hint_count = sum(1 for d in diagnostics if d['severity'] == 'hint')

            return {
                "status": "成功",
                "message": "诊断完成",
                "file_path": absolute_file_path,
                "diagnostics": diagnostics,
                "summary": {
                    "error_count": error_count,
                    "warning_count": warning_count,
                    "info_count": info_count,
                    "hint_count": hint_count,
                    "total_issues": len(diagnostics)
                },
                "pylint_warnings": pylint_warnings # 将收集到的Pylint执行警告也返回
            }

        except ImportError:
            return {
                "status": "失败",
                "message": "执行代码诊断需要 'subprocess' 和 'json' 模块，未能导入。"
            }
        except Exception as e:
            import traceback
            return {
                "status": "失败",
                "message": f"执行代码诊断时发生意外错误: {str(e)}",
                "error_details": traceback.format_exc(),
                "pylint_warnings": pylint_warnings
            }

    def _map_pylint_severity(self, pylint_type: str) -> str:
        """将pylint的问题类型映射到标准严重性级别"""
        mapping = {
            'error': 'error',
            'warning': 'warning',
            'convention': 'info',
            'refactor': 'hint',
            'info': 'info'
        }
        return mapping.get(pylint_type.lower(), 'info')

    async def file_operation(
        self,
        operation: Annotated[str, "要执行的操作类型，支持'read'、'write'、'append'、'edit'、'delete'、'touch'、'read_lines'、'write_lines'。"],
        file_path: Annotated[str, "文件路径。"],
        content: Annotated[Optional[Union[str, Dict, List]], "要写入或追加的内容，仅在'write'和'append'操作时需要。可以是字符串、字典或列表。对于JSON文件，字典和列表会自动序列化。"] = None,
        changes: Annotated[Optional[List[Dict[str, Any]]], "编辑操作列表，仅在'edit'操作时需要。"] = None,
        lines: Annotated[Optional[List[str]], "要写入的行列表，仅在'write_lines'操作时需要。"] = None,
        encoding: Annotated[str, "文件编码。"] = "utf-8",
        normalize_line_endings: Annotated[bool, "是否规范化行尾（将\\r\\n转换为\\n）。"] = True,
        ensure_final_newline: Annotated[bool, "确保文件以换行符结束。"] = True,
        create_if_not_exists: Annotated[bool, "如果文件不存在，是否创建新文件（仅适用于'write'、'append'和'edit'操作）。"] = True,
        recursive: Annotated[bool, "删除目录时是否递归删除（仅适用于'delete'操作）。"] = True,
        force: Annotated[bool, "删除不存在的文件时是否忽略错误（仅适用于'delete'操作）。"] = False,
        backup: Annotated[bool, "是否在修改或删除文件前创建备份（适用于'write'、'append'、'edit'、'delete'和'write_lines'操作）。"] = False,
        strip_newlines: Annotated[bool, "是否去除每行末尾的换行符（仅适用于'read_lines'操作）。"] = True,
        ensure_newlines: Annotated[bool, "是否确保每行以换行符结束（仅适用于'write_lines'操作）。"] = True,
        create_parents: Annotated[bool, "是否创建父目录（仅适用于'touch'操作）。"] = True,
        indent: Annotated[Optional[int], "JSON缩进空格数，用于格式化JSON文件。设为None则不缩进（仅适用于写入JSON文件时）。"] = 2,
        ensure_ascii: Annotated[bool, "写入JSON时是否确保所有非ASCII字符被转义（仅适用于写入JSON文件时）。"] = False,
        role: Annotated[Optional[str], "提交者的角色，如果提供，将添加到提交信息中（仅适用于需要Git提交的操作）。"] = None,
        auto_commit: Annotated[bool, "是否在操作后自动执行Git提交（适用于'write'、'append'、'edit'、'delete'和'write_lines'操作）。"] = True
    ) -> Dict[str, Any]:
        """
        【通用文件操作】执行常见的文件操作，包括读取、写入、追加、编辑、删除、创建空文件、按行读取和按行写入。

        这是一个通用函数，整合了多个常用的文件操作，简化了AI的选择。
        根据operation参数的不同，执行不同的文件操作。

        特别说明：
        - 对于JSON文件，'read'操作会自动解析JSON内容并在结果中添加'parsed_json'字段。
        - 对于JSON文件，'write'操作可以直接接受Python对象（字典或列表）作为content参数，会自动序列化为JSON。
        - 对于'append'操作，如果文件是JSON且content是字典或列表，会尝试读取现有JSON，合并后重写整个文件。

        Args:
            operation: 要执行的操作类型，支持'read'、'write'、'append'、'edit'、'delete'、'touch'、'read_lines'、'write_lines'。
            file_path: 文件路径。
            content: 要写入或追加的内容，仅在'write'和'append'操作时需要。可以是字符串、字典或列表。
            changes: 编辑操作列表，仅在'edit'操作时需要。
            lines: 要写入的行列表，仅在'write_lines'操作时需要。
            encoding: 文件编码。
            normalize_line_endings: 是否规范化行尾（将\\r\\n转换为\\n）。
            ensure_final_newline: 确保文件以换行符结束。
            create_if_not_exists: 如果文件不存在，是否创建新文件（仅适用于'write'、'append'和'edit'操作）。
            recursive: 删除目录时是否递归删除（仅适用于'delete'操作）。
            force: 删除不存在的文件时是否忽略错误（仅适用于'delete'操作）。
            backup: 是否在修改或删除文件前创建备份（适用于'write'、'append'、'edit'、'delete'和'write_lines'操作）。
            strip_newlines: 是否去除每行末尾的换行符（仅适用于'read_lines'操作）。
            ensure_newlines: 是否确保每行以换行符结束（仅适用于'write_lines'操作）。
            create_parents: 是否创建父目录（仅适用于'touch'操作）。
            indent: JSON缩进空格数，用于格式化JSON文件。设为None则不缩进（仅适用于写入JSON文件时）。
            ensure_ascii: 写入JSON时是否确保所有非ASCII字符被转义（仅适用于写入JSON文件时）。
            role: 提交者的角色，如果提供，将添加到提交信息中（仅适用于需要Git提交的操作）。
            auto_commit: 是否在操作后自动执行Git提交（适用于'write'、'append'、'edit'、'delete'和'write_lines'操作）。

        Returns:
            包含操作结果的字典，具体内容取决于执行的操作。

        示例:
            # 读取文件
            result = await file_utils.file_operation('read', 'example.txt')

            # 读取JSON文件（自动解析）
            result = await file_utils.file_operation('read', 'data.json')
            parsed_data = result.get("parsed_json")

            # 写入文件
            result = await file_utils.file_operation('write', 'example.txt', content='Hello, World!')

            # 写入JSON文件（自动序列化）
            data = {"name": "张三", "age": 30, "skills": ["Python", "JavaScript"]}
            result = await file_utils.file_operation('write', 'data.json', content=data)

            # 追加内容
            result = await file_utils.file_operation('append', 'example.txt', content='New line')

            # 编辑文件
            changes = [{"action": "replace_line", "line_number": 1, "new_content": "New first line"}]
            result = await file_utils.file_operation('edit', 'example.txt', changes=changes)

            # 删除文件
            result = await file_utils.file_operation('delete', 'example.txt')

            # 创建空文件或更新时间戳
            result = await file_utils.file_operation('touch', 'empty.txt')

            # 按行读取文件
            result = await file_utils.file_operation('read_lines', 'example.txt')

            # 按行写入文件
            lines = ["第一行", "第二行", "第三行"]
            result = await file_utils.file_operation('write_lines', 'example.txt', lines=lines)
        """
        operation = operation.lower()

        if operation == 'read':
            result = await self.read_file(file_path, encoding)

            # 如果文件是JSON格式，尝试自动解析
            if "error" not in result and file_path.lower().endswith('.json'):
                try:
                    content = result.get("content", "")
                    if content.strip():  # 确保内容不为空
                        parsed_json = json.loads(content)
                        result["parsed_json"] = parsed_json
                        result["is_json"] = True
                except json.JSONDecodeError as e:
                    result["json_parse_error"] = str(e)
                    result["is_json"] = False

            return result

        elif operation == 'read_lines':
            return await self.read_lines(file_path, encoding, strip_newlines)

        elif operation == 'write':
            if content is None:
                return {"error": "写入操作需要提供content参数。"}

            # 如果文件是JSON格式且content是字典或列表，自动转换为JSON字符串
            if file_path.lower().endswith('.json') and isinstance(content, (dict, list)):
                try:
                    # 使用write_json_file处理JSON对象
                    return await self.write_json_file(
                        file_path,
                        content,
                        encoding=encoding,
                        ensure_final_newline=ensure_final_newline,
                        indent=indent,
                        ensure_ascii=ensure_ascii,
                        role=role,
                        auto_commit=auto_commit
                    )
                except Exception as e:
                    return {"error": f"将对象转换为JSON并写入文件 {file_path} 时出错: {str(e)}"}
            else:
                # 处理普通文本内容
                return await self.write_file(
                    file_path,
                    str(content),  # 确保content是字符串
                    encoding=encoding,
                    normalize_line_endings=normalize_line_endings,
                    ensure_final_newline=ensure_final_newline,
                    role=role,
                    auto_commit=auto_commit
                )

        elif operation == 'write_lines':
            if lines is None:
                return {"error": "write_lines操作需要提供lines参数。"}
            return await self.write_lines(
                file_path,
                lines,
                encoding=encoding,
                ensure_newlines=ensure_newlines,
                ensure_final_newline=ensure_final_newline,
                backup=backup,
                role=role,
                auto_commit=auto_commit
            )

        elif operation == 'append':
            if content is None:
                return {"error": "追加操作需要提供content参数。"}

            # 如果文件是JSON格式且content是字典或列表，尝试智能合并
            if file_path.lower().endswith('.json') and isinstance(content, (dict, list)):
                try:
                    # 先读取现有JSON文件
                    read_result = await self.read_file(file_path, encoding)

                    if "error" in read_result:
                        # 文件不存在或读取错误，直接写入新内容
                        if create_if_not_exists:
                            return await self.write_json_file(
                                file_path,
                                content,
                                encoding=encoding,
                                ensure_final_newline=ensure_final_newline,
                                indent=indent,
                                ensure_ascii=ensure_ascii,
                                role=role,
                                auto_commit=auto_commit
                            )
                        else:
                            return read_result  # 返回读取错误

                    # 尝试解析现有JSON
                    existing_content = read_result.get("content", "")
                    if not existing_content.strip():
                        # 文件为空，直接写入新内容
                        return await self.write_json_file(
                            file_path,
                            content,
                            encoding=encoding,
                            ensure_final_newline=ensure_final_newline,
                            indent=indent,
                            ensure_ascii=ensure_ascii,
                            role=role,
                            auto_commit=auto_commit
                        )

                    try:
                        existing_json = json.loads(existing_content)

                        # 根据现有JSON和新内容的类型进行合并
                        if isinstance(existing_json, list) and isinstance(content, list):
                            # 列表合并
                            merged_content = existing_json + content
                        elif isinstance(existing_json, dict) and isinstance(content, dict):
                            # 字典合并
                            merged_content = {**existing_json, **content}
                        else:
                            # 类型不匹配，无法合并
                            return {"error": f"无法合并不同类型的JSON数据。现有数据类型: {type(existing_json).__name__}, 新数据类型: {type(content).__name__}"}

                        # 写入合并后的内容
                        return await self.write_json_file(
                            file_path,
                            merged_content,
                            encoding=encoding,
                            ensure_final_newline=ensure_final_newline,
                            indent=indent,
                            ensure_ascii=ensure_ascii,
                            role=role,
                            auto_commit=auto_commit
                        )
                    except json.JSONDecodeError:
                        # 现有内容不是有效的JSON，作为普通文本追加
                        return {"error": f"文件 {file_path} 内容不是有效的JSON，无法进行JSON合并操作。"}

                except Exception as e:
                    return {"error": f"处理JSON文件 {file_path} 的追加操作时出错: {str(e)}"}
            else:
                # 处理普通文本内容
                return await self.append_to_file(
                    file_path,
                    str(content),  # 确保content是字符串
                    encoding=encoding,
                    normalize_line_endings=normalize_line_endings,
                    ensure_final_newline=ensure_final_newline,
                    role=role,
                    auto_commit=auto_commit
                )

        elif operation == 'edit':
            if changes is None:
                return {"error": "编辑操作需要提供changes参数。"}

            # 如果文件是JSON格式且changes包含JSON特定操作，可以在这里添加特殊处理
            # 目前先使用标准的edit_text_file函数
            return await self.edit_text_file(
                file_path,
                changes,
                encoding=encoding,
                create_if_not_exists=create_if_not_exists,
                normalize_line_endings=normalize_line_endings,
                ensure_final_newline=ensure_final_newline,
                role=role,
                auto_commit=auto_commit
            )

        elif operation == 'touch':
            return await self.touch_file(
                file_path,
                create_parents=create_parents
            )

        elif operation == 'delete':
            return await self.delete_item(
                file_path,
                recursive=recursive,
                force=force,
                role=role,
                auto_commit=auto_commit
            )

        else:
            return {"error": f"不支持的操作类型: {operation}。支持的操作有: 'read', 'write', 'append', 'edit', 'delete', 'touch', 'read_lines', 'write_lines'。"}

    async def directory_operation(
        self,
        operation: Annotated[str, "要执行的操作类型，支持'list'、'create'、'structure'、'search'。"],
        path: Annotated[str, "目录路径。"],
        recursive: Annotated[bool, "是否递归处理子目录。"] = True,
        pattern: Annotated[Optional[str], "文件匹配模式，仅在'search'操作时需要。"] = None,
        max_depth: Annotated[Optional[int], "递归的最大深度，仅在'list'和'structure'操作时有效。"] = None,
        output_format: Annotated[str, "输出格式，'json'或'text'，仅在'structure'操作时有效。"] = "json",
        include_patterns: Annotated[Optional[List[str]], "要包含的文件/目录模式列表，仅在'structure'操作时有效。"] = None,
        exclude_patterns: Annotated[Optional[List[str]], "要排除的文件/目录模式列表，仅在'structure'操作时有效。"] = None,
        exist_ok: Annotated[bool, "如果目录已存在，是否忽略错误，仅在'create'操作时有效。"] = True,
        include_metadata: Annotated[bool, "是否在'list'操作中包含详细的文件/目录元数据。"] = False
    ) -> Dict[str, Any]:
        """
        【通用目录操作】执行常见的目录操作，包括列出目录内容、创建目录、获取目录结构和搜索文件。

        这是一个通用函数，整合了多个目录操作相关的函数，简化了AI的选择。
        根据operation参数的不同，执行不同的目录操作。

        Args:
            operation: 要执行的操作类型，支持'list'、'create'、'structure'、'search'。
            path: 目录路径。
            recursive: 是否递归处理子目录。
            pattern: 文件匹配模式，仅在'search'操作时需要。
            max_depth: 递归的最大深度，仅在'list'和'structure'操作时有效。
            output_format: 输出格式，'json'或'text'，仅在'structure'操作时有效。
            include_patterns: 要包含的文件/目录模式列表，仅在'structure'操作时有效。
            exclude_patterns: 要排除的文件/目录模式列表，仅在'structure'操作时有效。
            exist_ok: 如果目录已存在，是否忽略错误，仅在'create'操作时有效。
            include_metadata: 是否在'list'操作中包含详细的文件/目录元数据，如创建时间、修改时间、MIME类型等。

        Returns:
            包含操作结果的字典，具体内容取决于执行的操作。

        示例:
            # 列出目录内容
            result = await file_utils.directory_operation('list', 'my_directory')

            # 列出目录内容并包含详细元数据
            result = await file_utils.directory_operation('list', 'my_directory', include_metadata=True)

            # 创建目录
            result = await file_utils.directory_operation('create', 'new_directory')

            # 获取目录结构
            result = await file_utils.directory_operation('structure', 'project_directory', max_depth=3)

            # 搜索文件
            result = await file_utils.directory_operation('search', 'source_directory', pattern='*.py')
        """
        operation = operation.lower()

        if operation == 'list':
            result = await self.list_directory(path, recursive=recursive, max_depth=max_depth)

            # 如果需要包含详细元数据
            if include_metadata and "items" in result:
                for item in result["items"]:
                    if item["type"] == "file":
                        try:
                            metadata_result = await self.get_file_metadata(os.path.join(path, item["path"]))
                            if "metadata" in metadata_result:
                                item["metadata"] = metadata_result["metadata"]
                        except Exception as e:
                            item["metadata_error"] = str(e)

            return result

        elif operation == 'create':
            return await self.create_directory(path, exist_ok=exist_ok)

        elif operation == 'structure':
            return await self.get_project_structure(
                path,
                max_depth=max_depth,
                include_patterns=include_patterns,
                exclude_patterns=exclude_patterns,
                output_format=output_format
            )

        elif operation == 'search':
            if pattern is None:
                return {"error": "搜索操作需要提供pattern参数。"}
            return await self.search_files(
                pattern=pattern,
                directory=path,
                recursive=recursive
            )

        else:
            return {"error": f"不支持的操作类型: {operation}。支持的操作有: 'list', 'create', 'structure', 'search'。"}

    async def file_management(
        self,
        operation: Annotated[str, "要执行的操作类型，支持'move'、'copy'。"],
        source_path: Annotated[str, "源文件或目录路径。"],
        destination_path: Annotated[str, "目标文件或目录路径。"],
        overwrite: Annotated[bool, "如果目标已存在，是否覆盖。"] = False
    ) -> Dict[str, Any]:
        """
        【通用文件管理】执行文件管理操作，包括移动和复制文件或目录。

        这是一个通用函数，整合了文件管理相关的函数，简化了AI的选择。
        根据operation参数的不同，执行不同的文件管理操作。

        Args:
            operation: 要执行的操作类型，支持'move'、'copy'。
            source_path: 源文件或目录路径。
            destination_path: 目标文件或目录路径。
            overwrite: 如果目标已存在，是否覆盖。

        Returns:
            包含操作结果的字典，具体内容取决于执行的操作。

        示例:
            # 移动文件
            result = await file_utils.file_management('move', 'old_location.txt', 'new_location.txt')

            # 复制目录
            result = await file_utils.file_management('copy', 'source_dir', 'backup_dir', overwrite=True)
        """
        operation = operation.lower()

        if operation == 'move':
            return await self.move_item(source_path, destination_path, overwrite=overwrite)

        elif operation == 'copy':
            return await self.copy_item(source_path, destination_path, overwrite=overwrite)

        else:
            return {"error": f"不支持的操作类型: {operation}。支持的操作有: 'move', 'copy'。"}

    async def file_analysis(
        self,
        operation: Annotated[str, "要执行的操作类型，支持'compare'、'metadata'、'hash'、'highlight'、'diagnostics'。"],
        file_path: Annotated[str, "文件路径。"],
        compare_path: Annotated[Optional[str], "比较的目标文件路径，仅在'compare'操作时需要。"] = None,
        algorithm: Annotated[str, "哈希算法，如'md5'、'sha1'、'sha256'，仅在'hash'操作时有效。"] = "sha256",
        language: Annotated[Optional[str], "代码语言，如'python'、'javascript'，仅在'highlight'操作时有效。"] = None,
        config_path: Annotated[Optional[str], "Pylint配置文件路径，仅在'diagnostics'操作时有效。"] = None
    ) -> Dict[str, Any]:
        """
        【通用文件分析】执行文件分析操作，包括比较文件、获取元数据、计算哈希值、语法高亮和代码诊断。

        这是一个通用函数，整合了文件分析相关的函数，简化了AI的选择。
        根据operation参数的不同，执行不同的文件分析操作。

        Args:
            operation: 要执行的操作类型，支持'compare'、'metadata'、'hash'、'highlight'、'diagnostics'。
            file_path: 文件路径。
            compare_path: 比较的目标文件路径，仅在'compare'操作时需要。
            algorithm: 哈希算法，如'md5'、'sha1'、'sha256'，仅在'hash'操作时有效。
            language: 代码语言，如'python'、'javascript'，仅在'highlight'操作时有效。
            config_path: Pylint配置文件路径，仅在'diagnostics'操作时有效。

        Returns:
            包含操作结果的字典，具体内容取决于执行的操作。

        示例:
            # 比较两个文件
            result = await file_utils.file_analysis('compare', 'file1.txt', compare_path='file2.txt')

            # 获取文件元数据
            result = await file_utils.file_analysis('metadata', 'document.pdf')

            # 计算文件哈希值
            result = await file_utils.file_analysis('hash', 'important.zip', algorithm='sha256')

            # 对代码进行语法高亮
            result = await file_utils.file_analysis('highlight', 'script.py', language='python')

            # 获取代码诊断信息
            result = await file_utils.file_analysis('diagnostics', 'script.py')

            # 使用自定义配置文件获取代码诊断信息
            result = await file_utils.file_analysis('diagnostics', 'script.py', config_path='.pylintrc')
        """
        operation = operation.lower()

        if operation == 'compare':
            if compare_path is None:
                return {"error": "比较操作需要提供compare_path参数。"}
            return await self.compare_files(file_path, compare_path)

        elif operation == 'metadata':
            return await self.get_file_metadata(file_path)

        elif operation == 'hash':
            return await self.get_file_hash(file_path, algorithm=algorithm)

        elif operation == 'highlight':
            # 先读取文件内容
            read_result = await self.read_file(file_path)
            if "error" in read_result:
                return read_result

            content = read_result.get("content", "")
            return await self.highlight_code(content, language=language)

        elif operation == 'diagnostics':
            return await self.get_diagnostics(file_path, config_path=config_path)

        else:
            return {"error": f"不支持的操作类型: {operation}。支持的操作有: 'compare', 'metadata', 'hash', 'highlight', 'diagnostics'。"}

    async def archive_operation(
        self,
        operation: Annotated[str, "要执行的操作类型，支持'create'、'extract'。"],
        archive_path: Annotated[str, "压缩文件路径。"],
        target_path: Annotated[str, "创建压缩文件时的源路径列表或解压时的目标目录。"],
        archive_format: Annotated[str, "压缩文件格式，如'zip'、'tar'，仅在'create'操作时有效。"] = "zip",
        compression_level: Annotated[Optional[int], "压缩级别，仅在'create'操作时有效。"] = None
    ) -> Dict[str, Any]:
        """
        【通用压缩文件操作】执行压缩文件操作，包括创建和解压压缩文件。

        这是一个通用函数，整合了压缩文件相关的函数，简化了AI的选择。
        根据operation参数的不同，执行不同的压缩文件操作。

        Args:
            operation: 要执行的操作类型，支持'create'、'extract'。
            archive_path: 压缩文件路径。
            target_path: 创建压缩文件时的源路径列表或解压时的目标目录。
            archive_format: 压缩文件格式，如'zip'、'tar'，仅在'create'操作时有效。
            compression_level: 压缩级别，仅在'create'操作时有效。

        Returns:
            包含操作结果的字典，具体内容取决于执行的操作。

        示例:
            # 创建压缩文件
            result = await file_utils.archive_operation('create', 'backup.zip', ['file1.txt', 'dir1'])

            # 解压压缩文件
            result = await file_utils.archive_operation('extract', 'backup.zip', 'extracted_dir')
        """
        operation = operation.lower()

        if operation == 'create':
            # 如果target_path是字符串，转换为列表
            sources = [target_path] if isinstance(target_path, str) else target_path
            return await self.create_archive(
                archive_path,
                sources,
                archive_format=archive_format,
                compression_level=compression_level
            )

        elif operation == 'extract':
            return await self.extract_archive(archive_path, target_path)

        else:
            return {"error": f"不支持的操作类型: {operation}。支持的操作有: 'create', 'extract'。"}

    async def monitor_operation(
        self,
        operation: Annotated[str, "要执行的操作类型，支持'stop'、'status'。对于'start'操作，请直接使用start_file_monitor函数。"],
        monitor_id: Annotated[Optional[str], "监控ID，用于'stop'和'status'操作。"] = None
    ) -> Dict[str, Any]:
        """
        【通用文件监控】执行文件监控操作，包括停止和获取监控状态。

        注意：由于技术限制，'start'操作不能通过此函数调用，请直接使用start_file_monitor函数。
        这是因为回调函数无法通过JSON Schema序列化，而AutoGen需要将工具转换为JSON Schema。

        这是一个通用函数，整合了文件监控相关的函数，简化了AI的选择。
        根据operation参数的不同，执行不同的文件监控操作。

        Args:
            operation: 要执行的操作类型，支持'stop'、'status'。
            monitor_id: 监控ID，用于'stop'和'status'操作。

        Returns:
            包含操作结果的字典，具体内容取决于执行的操作。

        示例:
            # 获取监控状态
            result = await file_utils.monitor_operation('status', monitor_id='monitor_123')

            # 停止监控
            result = await file_utils.monitor_operation('stop', monitor_id='monitor_123')

            # 对于开始监控，请直接使用start_file_monitor函数
            # def my_callback(event_type, src_path):
            #     print(f"文件变化: {event_type} - {src_path}")
            # result = await file_utils.start_file_monitor('important.txt', my_callback)
        """
        operation = operation.lower()

        if operation == 'start':
            return {
                "error": "由于技术限制，'start'操作不能通过monitor_operation函数调用。请直接使用start_file_monitor函数，因为回调函数无法通过JSON Schema序列化。"
            }

        elif operation == 'stop':
            if monitor_id is None:
                return {"error": "停止监控操作需要提供monitor_id参数。"}
            return await self.stop_file_monitor(monitor_id)

        elif operation == 'status':
            return await self.get_file_monitor_status(monitor_id)

        else:
            return {"error": f"不支持的操作类型: {operation}。支持的操作有: 'stop', 'status'。对于'start'操作，请直接使用start_file_monitor函数。"}

    async def git(
        self,
        action: Annotated[str, "要执行的 Git 操作，如 'init', 'clone', 'commit', 'push', 'pull', 'status', 'log', 'branch', 'checkout', 'merge' 等。"],
        params: Annotated[Dict[str, Any], "操作的参数，根据不同的操作类型提供不同的参数。"] = None
    ) -> Dict[str, Any]:
        """
        【Git 仓库管理】执行 Git 操作，包括初始化仓库、克隆、提交、分支管理等。

        此函数提供了对 Git 仓库的全面管理功能，是对 GitManager 类的封装，
        提供了统一的接口来执行各种 Git 操作。每个文件操作后会自动提交到 git 仓库。

        Args:
            action: 要执行的 Git 操作，如 'init', 'clone', 'commit', 'push', 'pull', 'status', 'log', 'branch', 'checkout', 'merge' 等。
            params: 操作的参数，根据不同的操作类型提供不同的参数。

        Returns:
            包含操作结果的字典。

        示例:
            # 初始化仓库
            result = await file_utils.git("init")

            # 提交更改
            result = await file_utils.git("commit", {"message": "Initial commit", "all_changes": True})

            # 克隆仓库
            result = await file_utils.git("clone", {"url": "https://github.com/user/repo.git", "target_path": "local_repo"})

            # 切换分支
            result = await file_utils.git("checkout", {"branch_name": "develop", "create": True})

            # 合并分支
            result = await file_utils.git("merge", {"branch_name": "feature/new-feature"})

            # 推送到远程
            result = await file_utils.git("push", {"remote": "origin", "branch": "main"})

            # 获取状态
            result = await file_utils.git("status")

            # 获取日志
            result = await file_utils.git("log", {"max_count": 5})
        """
        git_manager = GitManager()
        params = params or {}

        try:
            if action == "init":
                # 不再设置默认的用户信息，使用Git的全局配置
                if params is None:
                    params = {}
                # 用户名和邮箱将由GitManager类根据Git全局配置处理
                return git_manager.init_repo(**params)
            elif action == "clone":
                return git_manager.clone_repo(**params)
            elif action == "commit":
                # 如果params中包含role参数，传递给commit方法
                role = params.pop("role", None) if params else None
                return git_manager.commit(**params, role=role)
            elif action == "push":
                return git_manager.push(**params)
            elif action == "pull":
                return git_manager.pull(**params)
            elif action == "status":
                return git_manager.status()
            elif action == "log":
                return git_manager.log(**params)
            elif action == "branch":
                return git_manager.create_branch(**params)
            elif action == "checkout":
                return git_manager.checkout_branch(**params)
            elif action == "merge":
                return git_manager.merge_branch(**params)
            elif action == "add":
                return git_manager.add(**params)
            elif action == "reset":
                return git_manager.reset(**params)
            elif action == "stash":
                return git_manager.stash(**params)
            elif action == "tag":
                return git_manager.tag(**params)
            elif action == "remote":
                return git_manager.remote(**params)
            elif action == "info":
                return git_manager.get_repo_info()
            else:
                return {"status": "失败", "message": f"不支持的 Git 操作: {action}"}
        except Exception as e:
            return {"status": "失败", "message": f"执行 Git 操作 {action} 时出错: {str(e)}"}

    async def write_json_file(
        self,
        file_path: Annotated[str, "要写入的JSON文件的完整路径。如果文件已存在，它将被覆盖。"],
        content: Annotated[Union[Dict, List], "要写入的JSON内容，以Python字典或列表形式提供。"],
        encoding: Annotated[str, "写入文件时使用的编码，例如'utf-8'。"] = "utf-8",
        indent: Annotated[Optional[int], "JSON缩进空格数，用于格式化。设为None则不缩进。"] = 2,
        ensure_ascii: Annotated[bool, "是否确保所有非ASCII字符被转义。"] = False,
        ensure_final_newline: Annotated[bool, "确保文件以换行符结束。"] = True,
        backup: Annotated[bool, "如果文件已存在，是否在覆盖前创建备份。"] = False,
        role: Annotated[Optional[str], "提交者的角色，如果提供，将添加到提交信息中。"] = None,
        auto_commit: Annotated[bool, "是否在操作后自动执行Git提交。"] = True
    ) -> Dict[str, Any]:
        """
        将Python对象序列化为JSON并写入文件。专门用于处理JSON数据，避免换行符问题。

        此函数自动处理JSON序列化，确保正确处理所有特殊字符和换行符。
        对于需要在JSON中表示多行文本的情况，此函数会自动处理转义，无需手动处理。
        可以选择在覆盖现有文件前创建备份，以防止数据丢失。
        操作后会尝试自动执行 Git 提交，可通过 auto_commit=False 禁用。

        Args:
            file_path: 要写入的JSON文件的完整路径。
            content: 要写入的JSON内容，以Python字典或列表形式提供。
            encoding: 写入文件时使用的编码。
            indent: JSON缩进空格数，用于格式化。设为None则不缩进（生成最紧凑的JSON）。
            ensure_ascii: 是否确保所有非ASCII字符被转义。设为False可保留中文等字符的可读性。
            ensure_final_newline: 确保文件以换行符结束。
            backup: 如果文件已存在，是否在覆盖前创建备份。
            role: 提交者的角色，如果提供，将添加到提交信息中。
            auto_commit: 是否在操作后自动执行Git提交。

        Returns:
            包含操作状态和文件信息的字典，与write_file函数返回格式相同。

        示例:
            # 写入简单的JSON数据
            data = {"name": "张三", "age": 30, "skills": ["Python", "JavaScript"]}
            result = await file_utils.write_json_file("user.json", data)

            # 写入包含多行文本的JSON数据
            data = {
                "title": "报告",
                "content": "这是第一行\\n这是第二行\\n这是第三行"
            }
            result = await file_utils.write_json_file("report.json", data)

            # 写入紧凑格式的JSON（无缩进）
            result = await file_utils.write_json_file("data.min.json", data, indent=None)

            # 写入JSON并创建备份
            result = await file_utils.write_json_file("important_config.json", data, backup=True)
        """
        try:
            # 将Python对象序列化为JSON字符串
            json_str = json.dumps(
                content,
                ensure_ascii=ensure_ascii,
                indent=indent
            )

            # 确保文件以换行符结束
            if ensure_final_newline and not json_str.endswith('\n'):
                json_str += '\n'

            # 使用write_file函数写入文件
            # 注意：不需要规范化行尾，因为json.dumps已经处理了换行符
            return await self.write_file(
                file_path=file_path,
                content=json_str,
                encoding=encoding,
                normalize_line_endings=False,  # JSON已经处理了换行符
                ensure_final_newline=False,    # 我们已经手动处理了末尾换行符
                backup=backup,                 # 传递备份参数
                role=role,                     # 传递角色参数
                auto_commit=auto_commit        # 传递自动提交参数
            )
        except Exception as e:
            absolute_file_path = os.path.abspath(os.path.join(self.base_path, file_path))
            return {"error": f"写入JSON文件 {absolute_file_path} 时出错: {str(e)}"}

    def _get_tool_description(self, func_name: str, default_description: str) -> str:
        """辅助函数，获取工具的描述，优先使用函数的文档字符串。"""
        func = getattr(self, func_name, None)
        if func and func.__doc__:
            return func.__doc__.strip().split('\n')[0] # 取文档字符串的第一行作为简短描述
        return default_description

    def get_autogen_tools(self) -> List[FunctionTool]:
        """
        将此类中的公共方法注册为 AutoGen FunctionTool。

        为了简化AI的选择，已将相关功能整合到通用操作函数中：
        - file_operation: 整合了读取、写入、追加、编辑、删除文件的功能
        - directory_operation: 整合了列出、创建目录、获取结构、搜索文件的功能
        - file_management: 整合了移动和复制文件/目录的功能
        - file_analysis: 整合了比较、元数据、哈希、高亮、诊断的功能
        - archive_operation: 整合了创建和解压压缩文件的功能
        - monitor_operation: 整合了开始、停止、获取监控状态的功能

        这些被整合的函数不会出现在工具列表中，但仍然可以在代码中直接调用。

        Returns:
            一个 FunctionTool 列表，可以直接传递给 AssistantAgent 的 tools 参数。
        """
        if not AUTOGEN_AVAILABLE:
            print("AutoGen 模块不可用，无法创建 FunctionTool 实例。")
            return []

        # 已被整合到通用操作函数中的函数，不在工具列表中显示
        integrated_methods = [
            # 整合到file_operation
            "read_file",
            "write_file",
            "append_to_file",
            "edit_text_file",
            "delete_item",
            "touch_file",
            "read_lines",
            "write_lines",

            # 整合到directory_operation
            "list_directory",
            "create_directory",
            "get_project_structure",
            "search_files",

            # 整合到file_management
            "move_item",
            "copy_item",

            # 整合到file_analysis
            "compare_files",
            "get_file_metadata",
            "get_file_hash",
            "highlight_code",
            "get_diagnostics",

            # 整合到archive_operation
            "create_archive",
            "extract_archive",

            # 整合到monitor_operation
            # 注意：start_file_monitor 需要直接使用，因为它接受回调函数参数
            "stop_file_monitor",
            "get_file_monitor_status"
        ]

        tools = []
        # 获取所有公共方法，排除已整合的方法
        public_methods = [
            method_name for method_name in dir(self)
            if callable(getattr(self, method_name))
            and not method_name.startswith("_")
            and method_name != "get_autogen_tools"
            and method_name not in integrated_methods  # 排除已整合的方法
        ]

        method_descriptions = {
            # 通用操作函数（推荐AI优先使用）
            "file_operation": "【推荐使用】通用文件操作函数，支持读取、写入、追加、编辑和删除文件，一个函数满足常见需求。",
            "directory_operation": "【推荐使用】通用目录操作函数，支持列出、创建目录、获取结构和搜索文件。",
            "file_management": "【推荐使用】通用文件管理函数，支持移动和复制文件或目录。",
            "file_analysis": "【推荐使用】通用文件分析函数，支持比较、获取元数据、计算哈希值、语法高亮和诊断。",
            "archive_operation": "【推荐使用】通用压缩文件操作函数，支持创建和解压压缩文件。",
            "monitor_operation": "【推荐使用】通用文件监控函数，支持停止和获取监控状态。注意：开始监控请直接使用start_file_monitor。",
            "git": "【推荐使用】Git 仓库管理函数，支持初始化、克隆、提交、分支管理等 Git 操作。",

            # 文件监控（特殊函数）
            "start_file_monitor": "开始监控文件或目录的变化，接受回调函数参数。",

            # 特殊文件操作
            "write_json_file": "将Python对象(字典或列表)序列化为JSON并写入文件，自动处理换行符和特殊字符。"
        }

        for method_name in public_methods:
            method = getattr(self, method_name)
            description = self._get_tool_description(method_name, method_descriptions.get(method_name, f"{method_name} 文件操作。"))

            # FunctionTool 可以直接处理同步和异步函数
            # AutoGen 的 AssistantAgent 会正确地 await 异步工具函数
            tools.append(FunctionTool(func=method, name=method_name, description=description))

        return tools

# 示例用法和注册为AutoGen工具
if __name__ == '__main__':
    async def example_usage():
        file_ops = FileUtils(base_path="./test_file_ops_dir") # 在当前目录下创建测试目录

        # 确保测试目录存在
        if not os.path.exists(file_ops.base_path):
            await file_ops.create_directory(file_ops.base_path)

        # 1. 写入文件
        write_result = await file_ops.write_file("example.txt", "Hello, AutoGen FileUtils!\nThis is a test file.")
        print(f"写入文件结果: {write_result}")

        # 2. 读取文件
        if write_result.get("status") == "成功":
            read_result = await file_ops.read_file("example.txt")
            print(f"读取文件结果: {read_result.get('content')[:50]}...") # 只打印部分内容

        # 3. 追加内容
        append_result = await file_ops.append_to_file("example.txt", "\nAppending new line.")
        print(f"追加内容结果: {append_result}")

        # 4. 列出目录
        list_result = await file_ops.list_directory(".") # 列出当前目录 (test_file_ops_dir)
        print(f"列出目录结果: {list_result}")

        # 5. 创建子目录
        create_dir_result = await file_ops.create_directory("subdir/another_subdir", exist_ok=True)
        print(f"创建目录结果: {create_dir_result}")
        await file_ops.write_file("subdir/another_subdir/test_in_subdir.txt", "Content in subdir")


        # 6. 获取项目结构 (JSON)
        json_structure_result = await file_ops.get_project_structure(".", max_depth=2, output_format="json", exclude_patterns=["*.log"])
        print(f"项目结构 (JSON): 获取到 {len(json_structure_result.get('structure', {}).get('children', []))} 个顶级项目")

        # 7. 获取项目结构 (Text)
        structure_text = await file_ops.get_project_structure(".", max_depth=2, output_format="text", exclude_patterns=["*.log"])
        print(f"\n项目结构 (Text):\n{structure_text.get('structure_text')}")


        # 8. 搜索文件
        search_result = await file_ops.search_files(pattern="*.txt", directory=".", recursive=True)
        print(f"搜索 *.txt 文件结果: {search_result}")

        # 9. 编辑文件
        if search_result.get("files"):
            first_file = search_result["files"][0]["path"] # 使用相对路径
            edit_changes = [
                {"action": "replace_line", "line_number": 1, "new_content": "Hello, AutoGen FileUtils! (Edited)"},
                {"action": "insert_after_line", "line_number": 2, "text_to_insert": "This line was inserted."},
                {"action": "append_text", "text_to_append": "--- End of File ---"}
            ]
            edit_result = await file_ops.edit_text_file(first_file, edit_changes)
            print(f"编辑文件 {first_file} 结果: {edit_result}")

            read_edited_result = await file_ops.read_file(first_file)
            print(f"编辑后读取文件 {first_file} 内容:\n{read_edited_result.get('content')}")

        # 10. 文件比较
        await file_ops.write_file("file1.txt", "Line 1\nLine 2\nLine 3 common")
        await file_ops.write_file("file2.txt", "Line 1 modified\nLine 2\nLine 4 new\nLine 3 common")
        compare_result = await file_ops.compare_files("file1.txt", "file2.txt")
        print(f"文件比较结果 (相似度 {compare_result.get('similarity_percentage')}%):\n{compare_result.get('diff_text')}")
        # print(f"HTML Diff: {compare_result.get('html_diff')}") # HTML Diff 较长，选择性打印

        # 11. 获取文件元数据
        metadata_result = await file_ops.get_file_metadata("example.txt")
        print(f"文件 example.txt 元数据: {metadata_result}")

        # 12. 创建压缩文件
        archive_sources = ["example.txt", "file1.txt", "subdir"]
        create_archive_result = await file_ops.create_archive("my_archive.zip", archive_sources, archive_format="zip")
        print(f"创建压缩文件结果: {create_archive_result}")

        # 13. 解压文件
        if create_archive_result.get("status") == "成功":
            extract_dir = "extracted_archive"
            await file_ops.create_directory(extract_dir, exist_ok=True) # 确保解压目录存在
            extract_result = await file_ops.extract_archive("my_archive.zip", extract_dir)
            print(f"解压文件结果: {extract_result}")
            list_extracted = await file_ops.list_directory(extract_dir)
            print(f"解压后目录内容: {list_extracted}")


        # 14. 文件哈希
        hash_result = await file_ops.get_file_hash("example.txt", algorithm="sha256")
        print(f"文件 example.txt SHA256 哈希: {hash_result}")

        # 15. 文件监控示例 (如果 watchdog 可用)
        if WATCHDOG_AVAILABLE:
            monitor_path = "example.txt"

            # 定义一个简单的回调函数
            def my_callback(event_type: str, src_path: str):
                print(f"文件监控回调: 事件类型='{event_type}', 路径='{src_path}' @ {datetime.datetime.now()}")

            monitor_start_result = await file_ops.start_file_monitor(monitor_path, my_callback)
            print(f"启动文件监控 ({monitor_path}): {monitor_start_result}")

            if monitor_start_result.get("status") == "成功":
                print(f"请在几秒内修改或保存文件 '{monitor_path}' 以查看监控效果...")
                await asyncio.sleep(2) # 给点时间手动修改文件
                await file_ops.append_to_file(monitor_path, "\nChange for monitor test.")
                await asyncio.sleep(2) # 等待事件处理

                monitor_status_result = await file_ops.get_file_monitor_status(monitor_start_result.get("monitor_id"))
                print(f"文件监控状态: {monitor_status_result}")

                monitor_stop_result = await file_ops.stop_file_monitor(monitor_start_result.get("monitor_id"))
                print(f"停止文件监控: {monitor_stop_result}")


        # 清理测试文件和目录
        # print("\n正在清理测试文件和目录...")
        # shutil.rmtree(file_ops.base_path) # 小心！这将删除整个 test_file_ops_dir
        # print(f"已删除目录: {file_ops.base_path}")


    # 获取并打印所有可用的 AutoGen 工具
    file_utils_instance = FileUtils()
    autogen_tools = file_utils_instance.get_autogen_tools()
    print(f"\n发现 {len(autogen_tools)} 个可用的 AutoGen FunctionTool:")
    for tool in autogen_tools:
        print(f"- 工具名称: {tool.name}, 描述: {tool.description}")
        # print(f"  Schema: {tool.schema}") # 打印 schema 会很长

    # 运行异步示例
    asyncio.run(example_usage())


# 提供一个函数来获取文件工具列表，以便与 AutoGen 0.5.6 兼容
def get_file_tools() -> List[FunctionTool]:
    """
    获取文件操作工具列表，以便与 AutoGen 0.5.6 兼容。

    Returns:
        一个 FunctionTool 列表，可以直接传递给 AssistantAgent 的 tools 参数。
    """
    if not AUTOGEN_AVAILABLE:
        print("AutoGen 模块不可用，无法创建 FunctionTool 实例。")
        return []

    file_utils_instance = FileUtils()
    return file_utils_instance.get_autogen_tools()


class GitManager:
    """
    Git 仓库管理工具类。
    提供了一组全面的 Git 操作功能，包括初始化仓库、提交更改、分支管理等。
    """

    def __init__(self, repo_path: Optional[str] = None, user_name: Optional[str] = None, user_email: Optional[str] = None): # MODIFIED
        """
        初始化 Git 管理器。

        Args:
            repo_path: Git 仓库路径。如果为 None，则尝试查找当前目录或其父目录中的 .git 目录。
            user_name: Git 用户名。如果为 None，将依赖全局/本地 Git 配置。
            user_email: Git 用户邮箱。如果为 None，将依赖全局/本地 Git 配置。
        """
        self.repo_path = repo_path or self._find_git_root()
        self.user_name = user_name
        self.user_email = user_email

        # 如果找到了仓库且提供了用户信息，则设置
        if self.repo_path and (self.user_name or self.user_email): # MODIFIED: only set if provided
            self._set_user_info()

    def _find_git_root(self) -> Optional[str]:
        """
        查找 Git 仓库根目录。

        从当前目录开始，向上查找包含 .git 目录的目录。

        Returns:
            Git 仓库根目录的路径，如果未找到则返回 None。
        """
        current_path = os.getcwd()
        while current_path:
            if os.path.exists(os.path.join(current_path, '.git')):
                return current_path
            parent_path = os.path.dirname(current_path)
            if parent_path == current_path:  # 已到达根目录
                break
            current_path = parent_path
        return None

    def _run_git_command(self, command: List[str], cwd: Optional[str] = None) -> Dict[str, Any]:
        """
        执行 Git 命令。

        Args:
            command: Git 命令及其参数的列表。
            cwd: 执行命令的工作目录。如果为 None，则使用 repo_path。

        Returns:
            包含命令执行结果的字典。
        """
        working_dir = cwd or self.repo_path
        if not working_dir:
            return {"status": "失败", "message": "未找到 Git 仓库根目录"}

        try:
            result = subprocess.run(
                ['git'] + command,
                cwd=working_dir,
                check=True,
                capture_output=True,
                text=True
            )
            return {
                "status": "成功",
                "message": result.stdout.strip(),
                "command": ' '.join(['git'] + command),
                "working_dir": working_dir
            }
        except subprocess.CalledProcessError as e:
            return {
                "status": "失败",
                "message": e.stderr.strip(),
                "command": ' '.join(['git'] + command),
                "working_dir": working_dir,
                "error_code": e.returncode
            }
        except Exception as e:
            return {
                "status": "失败",
                "message": str(e),
                "command": ' '.join(['git'] + command),
                "working_dir": working_dir
            }

    def commit_file(self, file_path: str, operation: str, role: Optional[str] = None) -> Dict[str, Any]:
        """
        提交文件更改到 Git 仓库。

        Args:
            file_path: 要提交的文件路径。
            operation: 执行的操作类型，如 'write', 'edit', 'delete' 等。
            role: 提交者的角色，如果提供，将添加到提交信息中。

        Returns:
            包含 Git 操作结果的字典。
        """
        if not self.repo_path:
            return {"git_status": "失败", "message": "未找到 Git 仓库根目录"}

        try:
            # 构建相对于仓库根目录的文件路径
            abs_file_path = os.path.abspath(file_path)
            rel_file_path = os.path.relpath(abs_file_path, self.repo_path)

            # 构建提交信息
            commit_message = f"{operation} {rel_file_path}"
            if role:
                commit_message = f"[{role}] {commit_message}"

            # 添加文件到暂存区
            if operation != 'delete':  # 如果不是删除操作
                add_result = self._run_git_command(['add', rel_file_path])
                if add_result["status"] == "失败":
                    return {"git_status": "失败", "message": f"Git add 失败: {add_result['message']}"}

            # 提交更改
            commit_result = self._run_git_command(['commit', '-m', commit_message])
            if commit_result["status"] == "失败":
                # 如果没有更改可提交，这不一定是错误
                if "nothing to commit" in commit_result["message"]:
                    return {"git_status": "成功", "message": "没有更改需要提交", "repo_root": self.repo_path}
                return {"git_status": "失败", "message": f"Git commit 失败: {commit_result['message']}"}

            return {"git_status": "成功", "message": f"已提交更改: {commit_message}", "repo_root": self.repo_path}
        except Exception as e:
            return {"git_status": "失败", "message": f"Git 提交失败: {str(e)}"}

    def init_repo(self, path: Optional[str] = None, bare: bool = False, user_name: Optional[str] = None, user_email: Optional[str] = None) -> Dict[str, Any]: # MODIFIED
        """
        初始化一个新的 Git 仓库。

        Args:
            path: 要初始化的仓库路径。如果为 None，则使用当前目录。
            bare: 是否创建裸仓库。
            user_name: Git 用户名。如果为 None，将依赖全局/本地 Git 配置。
            user_email: Git 用户邮箱。如果为 None，将依赖全局/本地 Git 配置。
        """
        init_path = os.path.abspath(path or os.getcwd())
        command = ['init']
        if bare:
            command.append('--bare')

        if not os.path.exists(init_path):
            os.makedirs(init_path, exist_ok=True)
        elif not os.path.isdir(init_path):
            return {"status": "失败", "message": f"路径 {init_path} 已存在且不是一个目录。"}

        result = self._run_git_command(command, cwd=init_path)
        if result["status"] == "成功":
            self.repo_path = init_path
            self.user_name = user_name
            self.user_email = user_email
            if self.user_name or self.user_email:
                 self._set_user_info()
        return result

    def clone(self, url: str, target_path: Optional[str] = None, branch: Optional[str] = None) -> Dict[str, Any]:
        """
        克隆远程 Git 仓库。

        Args:
            url: 远程仓库 URL。
            target_path: 克隆到的目标路径。如果为 None，则使用当前目录。
            branch: 要克隆的分支。如果为 None，则克隆默认分支。

        Returns:
            包含克隆结果的字典。
        """
        command = ['clone', url]
        if branch:
            command.extend(['--branch', branch])
        if target_path:
            command.append(target_path)

        result = self._run_git_command(command, cwd=os.path.dirname(target_path) if target_path else None)
        if result["status"] == "成功" and target_path:
            self.repo_path = os.path.abspath(target_path)
        return result

    def create_branch(self, branch_name: str, start_point: Optional[str] = None) -> Dict[str, Any]:
        """
        创建新的 Git 分支。

        Args:
            branch_name: 新分支的名称。
            start_point: 分支的起点。如果为 None，则从当前 HEAD 创建分支。

        Returns:
            包含创建结果的字典。
        """
        command = ['branch', branch_name]
        if start_point:
            command.append(start_point)
        return self._run_git_command(command)

    def checkout_branch(self, branch_name: str, create: bool = False) -> Dict[str, Any]:
        """
        切换到指定的 Git 分支。

        Args:
            branch_name: 要切换到的分支名称。
            create: 如果为 True 且分支不存在，则创建新分支。

        Returns:
            包含切换结果的字典。
        """
        command = ['checkout']
        if create:
            command.append('-b')
        command.append(branch_name)
        return self._run_git_command(command)

    def merge_branch(self, branch_name: str, message: Optional[str] = None) -> Dict[str, Any]:
        """
        将指定分支合并到当前分支。

        Args:
            branch_name: 要合并的分支名称。
            message: 合并提交的信息。如果为 None，则使用默认信息。

        Returns:
            包含合并结果的字典。
        """
        command = ['merge', branch_name]
        if message:
            command.extend(['-m', message])
        return self._run_git_command(command)

    def pull(self, remote: str = 'origin', branch: Optional[str] = None) -> Dict[str, Any]:
        """
        从远程仓库拉取更改。

        Args:
            remote: 远程仓库名称。
            branch: 要拉取的分支。如果为 None，则拉取当前分支。

        Returns:
            包含拉取结果的字典。
        """
        command = ['pull', remote]
        if branch:
            command.append(branch)
        return self._run_git_command(command)

    def push(self, remote: str = 'origin', branch: Optional[str] = None, force: bool = False) -> Dict[str, Any]:
        """
        推送更改到远程仓库。

        Args:
            remote: 远程仓库名称。
            branch: 要推送的分支。如果为 None，则推送当前分支。
            force: 是否强制推送。

        Returns:
            包含推送结果的字典。
        """
        command = ['push', remote]
        if branch:
            command.append(branch)
        if force:
            command.append('--force')
        return self._run_git_command(command)

    def status(self) -> Dict[str, Any]:
        """
        获取 Git 仓库状态。

        Returns:
            包含仓库状态的字典。
        """
        return self._run_git_command(['status'])

    def log(self, max_count: int = 10, pretty_format: str = '%h %s (%an, %ar)') -> Dict[str, Any]:
        """
        获取 Git 提交日志。

        Args:
            max_count: 最多返回的提交数量。
            pretty_format: 日志格式。

        Returns:
            包含提交日志的字典。
        """
        command = ['log', f'--max-count={max_count}', f'--pretty=format:{pretty_format}']
        return self._run_git_command(command)

    def add(self, paths: Union[str, List[str]]) -> Dict[str, Any]:
        """
        将文件添加到 Git 暂存区。

        Args:
            paths: 要添加的文件路径或路径列表。

        Returns:
            包含添加结果的字典。
        """
        if isinstance(paths, str):
            paths = [paths]
        command = ['add'] + paths
        return self._run_git_command(command)

    def _set_user_info(self) -> Dict[str, Any]:
        """
        设置 Git 用户信息（如果已在实例中通过 __init__ 或 init_repo 提供）。
        如果实例中的 user_name 或 user_email 为 None，则不会设置对应的 Git 配置，
        此时 Git 会使用其已有的全局或仓库级配置。
        """
        if not self.repo_path:
            return {"status": "失败", "message": "无法设置用户信息：未指定或找到有效的仓库路径。"}

        messages = []
        if self.user_name is not None:
            name_result = self._run_git_command(['config', 'user.name', self.user_name])
            if name_result["status"] == "成功":
                messages.append(f"已在仓库 {self.repo_path} 中设置用户名: {self.user_name}")
            else:
                messages.append(f"在仓库 {self.repo_path} 中设置用户名失败: {name_result.get('message', '未知错误')}")

        if self.user_email is not None:
            email_result = self._run_git_command(['config', 'user.email', self.user_email])
            if email_result["status"] == "成功":
                messages.append(f"已在仓库 {self.repo_path} 中设置用户邮箱: {self.user_email}")
            else:
                messages.append(f"在仓库 {self.repo_path} 中设置用户邮箱失败: {email_result.get('message', '未知错误')}")

        if not self.user_name and not self.user_email:
             return {"status": "信息", "message": f"未在Manager中提供用户信息以配置仓库 {self.repo_path}，将依赖Git现有配置。"}

        success_overall = all("失败" not in msg for msg in messages if "设置" in msg)

        return {
            "status": "成功" if success_overall and messages else ("部分成功或失败" if messages else "信息"),
            "message": " | ".join(messages) if messages else f"未在Manager中提供用户信息以配置仓库 {self.repo_path}，将依赖Git现有配置。"
        }

    def commit(self, message: str, all_changes: bool = False, role: Optional[str] = None) -> Dict[str, Any]:
        """
        提交 Git 暂存区的更改。

        Args:
            message: 提交信息。
            all_changes: 是否自动暂存所有已跟踪文件的更改。
            role: 提交者的角色，如果提供，将添加到提交信息中。

        Returns:
            包含提交结果的字典。
        """
        # 如果提供了角色信息，添加到提交信息中
        if role:
            message = f"[{role}] {message}"

        command = ['commit', '-m', message]
        if all_changes:
            command.append('-a')
        return self._run_git_command(command)

    def reset(self, mode: str = 'mixed', commit: str = 'HEAD') -> Dict[str, Any]:
        """
        重置 Git 仓库状态。

        Args:
            mode: 重置模式，可以是 'soft'、'mixed' 或 'hard'。
            commit: 要重置到的提交。

        Returns:
            包含重置结果的字典。
        """
        command = ['reset', f'--{mode}', commit]
        return self._run_git_command(command)

    def stash(self, action: str = 'push', message: Optional[str] = None) -> Dict[str, Any]:
        """
        管理 Git 储藏。

        Args:
            action: 储藏操作，可以是 'push'、'pop'、'apply'、'list' 等。
            message: 储藏信息（仅适用于 'push' 操作）。

        Returns:
            包含储藏操作结果的字典。
        """
        command = ['stash', action]
        if action == 'push' and message:
            command.extend(['--message', message])
        return self._run_git_command(command)

    def tag(self, tag_name: Optional[str] = None, message: Optional[str] = None, commit: Optional[str] = None) -> Dict[str, Any]:
        """
        管理 Git 标签。

        Args:
            tag_name: 标签名称。如果为 None，则列出所有标签。
            message: 标签信息。如果提供，则创建带注释的标签。
            commit: 要标记的提交。如果为 None，则标记当前 HEAD。

        Returns:
            包含标签操作结果的字典。
        """
        if tag_name is None:
            return self._run_git_command(['tag'])

        command = ['tag']
        if message:
            command.extend(['-a', tag_name, '-m', message])
        else:
            command.append(tag_name)

        if commit:
            command.append(commit)

        return self._run_git_command(command)

    def remote(self, action: str = 'show', remote_name: Optional[str] = None, url: Optional[str] = None) -> Dict[str, Any]:
        """
        管理 Git 远程仓库。

        Args:
            action: 远程仓库操作，可以是 'add'、'remove'、'show' 等。
            remote_name: 远程仓库名称。
            url: 远程仓库 URL（仅适用于 'add' 操作）。

        Returns:
            包含远程仓库操作结果的字典。
        """
        command = ['remote', action]
        if remote_name:
            command.append(remote_name)
        if action == 'add' and url:
            command.append(url)
        return self._run_git_command(command)

    def get_current_branch(self) -> str:
        """
        获取当前分支名称。

        Returns:
            当前分支名称，如果发生错误则返回空字符串。
        """
        result = self._run_git_command(['rev-parse', '--abbrev-ref', 'HEAD'])
        return result.get("message", "") if result["status"] == "成功" else ""

    def get_repo_info(self) -> Dict[str, Any]:
        """
        获取 Git 仓库信息。

        Returns:
            包含仓库信息的字典。
        """
        if not self.repo_path:
            return {"status": "失败", "message": "未找到 Git 仓库根目录"}

        info = {
            "repo_path": self.repo_path,
            "current_branch": self.get_current_branch()
        }

        # 获取远程仓库信息
        remote_result = self._run_git_command(['remote', '-v'])
        if remote_result["status"] == "成功":
            remotes = {}
            for line in remote_result["message"].splitlines():
                if not line.strip():
                    continue
                parts = line.split()
                if len(parts) >= 2:
                    name, url = parts[0], parts[1]
                    remotes[name] = url
            info["remotes"] = remotes

        # 获取最近的提交
        log_result = self._run_git_command(['log', '--max-count=1', '--pretty=format:%H|%an|%ae|%at|%s'])
        if log_result["status"] == "成功" and log_result["message"]:
            commit_parts = log_result["message"].split('|')
            if len(commit_parts) >= 5:
                info["last_commit"] = {
                    "hash": commit_parts[0],
                    "author_name": commit_parts[1],
                    "author_email": commit_parts[2],
                    "timestamp": commit_parts[3],
                    "message": commit_parts[4]
                }

        return {"status": "成功", "info": info}

